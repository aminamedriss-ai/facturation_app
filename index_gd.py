import requests
import streamlit as st
import pandas as pd
import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
import bcrypt
import json
import os
import hashlib
import base64
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
import calendar
import math
import numpy as np
import io
from decimal import Decimal
from pathlib import Path
from rapidfuzz import process, fuzz
from reportlab.platypus import Image
from supabase import create_client, Client
from PIL import Image  
import pytesseract
import easyocr
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.errors import HttpError
import os, pickle
import webbrowser
import threading
import time
import pkg_resources
from math import floor
import unicodedata
import re
import difflib
from reportlab.platypus import Image, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
import os
from datetime import datetime
# 📷 Afficher un logo
st.set_page_config(
    page_title="Gestion de la Facturation",
    page_icon="logo2.png",  # chemin local ou URL
    layout="wide"
)
st.markdown("""
<style>
/* 🖼 Logo centré */
.logo-container {
    display: flex;
    justify-content: center;
    align-items: center;
    margin-bottom: 20px;
}

.logo {
    width: 200px;
    height: auto;
}

/* 🌈 Arrière-plan personnalisé + forcer mode sombre */
html, body, .stApp {
    background: #1d2e4e !important;
    font-family: 'Segoe UI', sans-serif;
    color-scheme: dark !important; /* Empêche l'inversion automatique */
    color: white !important;
}

/* 🖍️ Titre centré et coloré */
.main > div > div > div > div > h1 {
    text-align: center;
    color: #00796B !important;
}

/* 🧼 Nettoyage des bordures Streamlit */
.css-18e3th9 {
    padding: 1rem 0.5rem;
}

/* 🎨 Sidebar */
section[data-testid="stSidebar"] {
    background-color: #1f3763 !important;
    color: white !important;
}

section[data-testid="stSidebar"] .css-1v3fvcr {
    color: white !important;
}

/* 🌈 Titres dans la sidebar */
section[data-testid="stSidebar"] h1, 
section[data-testid="stSidebar"] h2, 
section[data-testid="stSidebar"] h3 {
    color: #e01b36 !important;
}

/* 🎨 Barre supérieure */
header[data-testid="stHeader"] {
    background-color: #06dbae !important;
    color: white !important;
}

/* 🧪 Supprimer la transparence */
header[data-testid="stHeader"]::before {
    content: "";
    background: none !important;
}

/* 📱 Correction mobile : forcer couleurs partout */
h1, h2, h3, p, span, label {
    color: white !important;
}

/* 🔵 Boutons bleu foncé forcés */
.stButton button {
    background-color: #2b2c36 !important; /* Bleu foncé */
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.5rem 1rem !important;
    font-weight: bold !important;
    -webkit-appearance: none !important; /* Évite style par défaut mobile */
    appearance: none !important;
}

.stButton button:hover {
    background-color: #43444e !important; /* Bleu plus clair au survol */
    color: white !important;
}
</style>
""", unsafe_allow_html=True)



# 🖼️ Ajouter un logo (remplacer "logo.png" par ton fichier ou une URL)
with open("logo.png", "rb") as image_file:
    encoded = base64.b64encode(image_file.read()).decode()

st.markdown(
    f"""
    <div class="logo-container">
        <img class="logo" src="data:image/png;base64,{encoded}">
    </div>
    """,
    unsafe_allow_html=True
)
st.markdown(
    "<h1 style='text-align: center;'>Bienvenue sur l'application de calcul des factures</h1>",
    unsafe_allow_html=True
)
def nettoyer_colonne(df, col):
    series = (
        df[col]
        .astype(str)
        .str.replace(r"[^\d,.-]", "", regex=True)
        .str.replace(",", ".", regex=False)
        .replace("", "0")
    )

    # Vérifie quelles valeurs ne sont pas convertibles
    erreurs = []
    for val in series.unique():
        try:
            float(val)
        except ValueError:
            erreurs.append(val)

    if erreurs:
        print(f"⚠️ Colonne '{col}' contient des valeurs non convertibles : {erreurs[:20]}")

    # Conversion sécurisée
    return pd.to_numeric(series, errors="coerce").fillna(0.0)




def calcul_base(row):
    salaire = Decimal(str(row["Salaire de base calcule"]))
    prime = Decimal(str(row["Prime mensuelle (Barème) (DZD)"]))
    panier = Decimal(str(row["Indemnité de panier (DZD)"]))
    transport  = Decimal(str(row["Indemnité de transport (DZD)"]))

    return floor((salaire + prime) - ((salaire + prime) * Decimal("0.09")) + (panier+transport))
def get_valeur(col_base, col_nouveau):
    if col_nouveau in df_client.columns:
        return nettoyer_colonne(df_client, col_nouveau).where(
            lambda x: x != 0, nettoyer_colonne(df_client, col_base)
        )
    else:
        return nettoyer_colonne(df_client, col_base)


def normalize_text(s):
    """Normalise pour comparaison: enlever diacritiques, NBSP, espaces multiples, lower."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def fetch_all_data(table_name, batch_size=1000):
    all_data = []
    last_id = 0
    
    while True:
        response = (
            supabase.table(table_name)
            .select("*")
            .order("id")
            .gte("id", last_id + 1)
            .limit(batch_size)
            .execute()
        )
        
        if not response.data:
            break
        
        all_data.extend(response.data)
        
        # mettre à jour le dernier id récupéré
        last_id = max(item["id"] for item in response.data)
        
        if len(response.data) < batch_size:
            break
    
    return pd.DataFrame(all_data)

def trouver_client_robuste(client_name, df, debug=False):
    """
    Filtrage tolérant pour retrouver les lignes d'un client dans df.
    Stratégies (en cascade):
      1) égalité normalisée
      2) contenant (normalized contains)
      3) fuzzy close match (difflib) sur la liste d'établissements uniques
    Retourne df_filtré (copie). Si debug=True, renvoie aussi info dict.
    """
    if df is None or df.empty:
        return (pd.DataFrame(), {"reason": "empty_df"}) if debug else pd.DataFrame()

    df2 = df.copy()
    df2["Etablissement_norm"] = df2["Etablissement"].astype(str).apply(normalize_text)
    client_norm = normalize_text(client_name)

    # 1) égalité normalisée
    mask_eq = df2["Etablissement_norm"] == client_norm
    df_eq = df2[mask_eq].copy()
    if not df_eq.empty:
        if debug:
            return df_eq, {"method": "exact_norm", "count": len(df_eq)}
        return df_eq

    # 2) contains normalisé (client substring of Etab_norm or vice-versa)
    mask_contains = df2["Etablissement_norm"].str.contains(re.escape(client_norm), na=False)
    df_contains = df2[mask_contains].copy()
    if not df_contains.empty:
        if debug:
            return df_contains, {"method": "contains_etab", "count": len(df_contains)}
        return df_contains

    # also try client contained in etab or etab contained in client (reverse)
    mask_rev = df2["Etablissement_norm"].apply(lambda x: client_norm in x or x in client_norm)
    df_rev = df2[mask_rev].copy()
    if not df_rev.empty:
        if debug:
            return df_rev, {"method": "contains_rev", "count": len(df_rev)}
        return df_rev

    # 3) fuzzy matching on unique values
    uniques = sorted(df2["Etablissement_norm"].dropna().unique().tolist())
    close = difflib.get_close_matches(client_norm, uniques, n=3, cutoff=0.7)
    if close:
        # prendre le premier best match
        best = close[0]
        df_close = df2[df2["Etablissement_norm"] == best].copy()
        if debug:
            return df_close, {"method": "fuzzy", "match": best, "close_candidates": close, "count": len(df_close)}
        return df_close

    # Aucun résultat
    if debug:
        # retourner un petit aperçu des candidats potentiels pour aider
        sample_uniques = uniques[:30]
        return pd.DataFrame(), {"method": "no_match", "sample_candidates": sample_uniques}

    return pd.DataFrame()


def calcul_joursstc_ouvres(row):
    val = row["Date du dernier jour travaillé"]
    nbr_jours_stc = row["Nbr jours STC (jours)"]
    
    date_debut = pd.to_datetime(val, dayfirst=True, errors="coerce")
    if pd.isna(date_debut) or nbr_jours_stc == 0:
        return 0

    start = (date_debut + pd.Timedelta(days=1)).normalize()
    end = start + pd.Timedelta(days=nbr_jours_stc - 1)

    # Générer toutes les dates de la période
    toutes_les_dates = pd.date_range(start=start, end=end, freq="D")
    
    # Filtrer les jours ouvrés (exclure vendredi=4 et samedi=5)
    jours_ouvres = sum(1 for date in toutes_les_dates if date.weekday() not in [4, 5])
    
    return jours_ouvres



def authenticate_drive():
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    service = build("drive", "v3", credentials=creds)

    try:
        # Test: lister quelques fichiers accessibles
        results = service.files().list(pageSize=5, fields="files(id, name)").execute()
        # st.success("✅ Connexion Drive réussie")
        # st.write(results.get("files", []))
    except Exception as e:
        st.error(f"❌ Erreur connexion Drive: {e}")

    return service



def get_or_create_folder(service, folder_name, parent_id=None, drive_id=None):
    """
    Vérifie si un dossier existe dans Google Drive (y compris Drive partagé), sinon le crée.
    Retourne l'ID du dossier.
    """
    # Requête pour chercher le dossier
    query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and trashed=false"
    if parent_id:
        query += f" and '{parent_id}' in parents"

    results = service.files().list(
        q=query,
        fields="files(id, name)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    items = results.get("files", [])

    if items:
        return items[0]["id"]  # ✅ Dossier trouvé

    # Sinon, création du dossier
    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder"
    }
    if parent_id:
        metadata["parents"] = [parent_id]
    if drive_id:  # obligatoire pour Drive partagé si parent_id est la racine
        metadata["driveId"] = drive_id

    folder = service.files().create(
        body=metadata,
        fields="id",
        supportsAllDrives=True
    ).execute()
    return folder["id"]



def upload_to_drive(file_path, client_name, root_folder_id=None, drive_id=None):
    service = authenticate_drive()

    # 1️⃣ Vérifier/créer le dossier client
    folder_id = get_or_create_folder(service, client_name, parent_id=root_folder_id, drive_id=drive_id)

    # Nom du fichier (nom local)
    file_name = os.path.basename(file_path)

    # 2️⃣ Vérifier si un fichier avec le même nom existe déjà
    query = f"name='{file_name}' and '{folder_id}' in parents and trashed=false"
    results = service.files().list(
        q=query,
        fields="files(id, name)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    existing_files = results.get("files", [])

    if existing_files:
        file_id = existing_files[0]["id"]
        print(f"♻️ Mise à jour du fichier existant : {file_name} ({file_id})")

        media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        file = service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True
        ).execute()
    else:
        file_metadata = {"name": file_name, "parents": [folder_id]}
        if drive_id:  # utile si on est dans un Drive partagé
            file_metadata["driveId"] = drive_id

        media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        try:
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields="id",
                supportsAllDrives=True
            ).execute()
        except HttpError as e:
            st.error("⚠️ Erreur API Google Drive")
            st.code(e.content.decode("utf-8") if hasattr(e, "content") else str(e))
            raise

    print(f"✅ Fichier disponible dans {client_name} : {file_name} ({file['id']})")
    return file["id"]
def create_recap_sheet(wb: Workbook,
                       df_pivot,
                       df_travel_pivot,
                       df_allowance_pivot,
                       taux_data,
                       devise="USD",
                       mois_ordre=None,
                       phone_gross_dzd=24000,
                       tax_pct_line2="5%",   # si tu veux afficher un % spécifique ligne 2 (None = vide)
                       apply_special_tax_pct=5,  # ligne 3 (tu avais demandé 5%)
                       tva_pct=19):
    """
    Crée une feuille "Overview" en tête du workbook wb avec 4 tableaux :
    1) Cost salary (Facture HT par employé x mois + totaux, TVA, taux, facture en devise)
    2) Travel expenses (M et C par employé x mois + totaux, taxe 5%, TVA, taux, en devise)
    3) Allowance (même structure que travel)
    4) Phone package 7-lignes comme spécifié.
    """

    # --- paramètres par défaut des mois ---
    if mois_ordre is None:
        mois_ordre = ["Janvier","Février","Mars","Avril","Mai","Juin",
                      "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # styles
    header_font = Font(bold=True, size=11)
    bold = Font(bold=True)
    normal_font_white = Font(size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    phone_gross_row = None
   
    font_white_bold = Font(color="FFFFFF", bold=True)


    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_sky_blue = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")   # Bleu ciel
    fill_dark_blue = PatternFill(start_color="284052", end_color="284052", fill_type="solid")  # Bleu un peu plus foncé

    # 📌 Dictionnaire des couleurs par libellé
    style_map = {
        "Invoice before VAT": fill_red,
        "Invoice including VAT": fill_sky_blue,
        "Taxe": fill_sky_blue,
        "change rate": fill_sky_blue,
        "USD equivalent before VAT": fill_dark_blue,
        "USD equivalent including VAT": fill_dark_blue,
        "Total Invoice before VAT (USD)": fill_dark_blue,
        "Total Invoice including VAT (USD)": fill_dark_blue
    }
    # créer la feuille en première position (supprime si existe)
    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        wb.remove(ws)
    ws = wb.create_sheet("Overview", 0)
    # 📏 Ajuster la largeur des colonnes
    # 📏 Ajuster la largeur des colonnes
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]:
        ws.column_dimensions[col].width = 30



    # helper pour récupérer colonnes par préfixe
    def monthly_cols(df, prefix):
        cols = []
        for m in mois_ordre:
            col = f"{prefix}_{m}"
            if col in df.columns:
                cols.append(col)
            else:
                # si manque, ajouter colonne virtuelle en 0 (gestion plus bas)
                cols.append(col)
        return cols

    # helper: safe fetch series (0 if missing)
    def get_series_safe(pivot_df, col):
        if col in pivot_df.columns:
            return pivot_df[col].fillna(0)
        else:
            # build zeros series with index pivot_df.index
            import pandas as pd
            return pd.Series(0, index=pivot_df.index)

    # --- Préparer listes d'employés (index) et mapping ---
    # On prend les employés présents dans df_pivot si possible sinon dans df_travel_pivot sinon df_allowance_pivot
    if hasattr(df_pivot, "index") and len(df_pivot) > 0:
        employees = df_pivot.index if isinstance(df_pivot.index, (list,)) else df_pivot.reset_index()["N°"]
        df_emp_index = df_pivot.reset_index()
    elif hasattr(df_travel_pivot, "index") and len(df_travel_pivot) > 0:
        df_emp_index = df_travel_pivot.reset_index()
    elif hasattr(df_allowance_pivot, "index") and len(df_allowance_pivot) > 0:
        df_emp_index = df_allowance_pivot.reset_index()
    else:
        # fallback si aucune data
        df_emp_index = df_pivot.reset_index() if hasattr(df_pivot, "reset_index") else df_pivot

    # attempt to obtain a list of employees with 'N°' and 'Nom' columns
    if "N°" in df_emp_index.columns:
        emp_ids = df_emp_index["N°"].unique().tolist()
    else:
        # try to use index values
        emp_ids = df_emp_index.index.astype(str).tolist()

    # build per-employee rows for table 1 (Facture HT)
    # identify monthly Facture HT cols
    facture_cols = monthly_cols(df_pivot, "Facture HT")
    # Travel monthly columns M and C
    travel_m_cols = monthly_cols(df_travel_pivot, "Travel expenses M segment")
    travel_c_cols = monthly_cols(df_travel_pivot, "Travel expenses C segment")
    # Allowance monthly columns
    allow_m_cols = monthly_cols(df_allowance_pivot, "Allowance M segment")
    allow_c_cols = monthly_cols(df_allowance_pivot, "Allowance C segment")

    # Prepare a DataFrame-like structure for writing:
    # For pivot DataFrames we assume they have 'N°' as column (reset_index)
    df_piv = df_pivot.reset_index() if "N°" not in df_pivot.columns else df_pivot.copy()
    df_pv_travel = df_travel_pivot.reset_index() if "N°" not in df_travel_pivot.columns else df_travel_pivot.copy()
    df_pv_allow = df_allowance_pivot.reset_index() if "N°" not in df_allowance_pivot.columns else df_allowance_pivot.copy()

    # ensure N° exists
    if "N°" not in df_piv.columns:
        if "index" in df_piv.columns:
            df_piv = df_piv.rename(columns={"index": "N°"})
    if "N°" not in df_pv_travel.columns:
        if "index" in df_pv_travel.columns:
            df_pv_travel = df_pv_travel.rename(columns={"index": "N°"})
    if "N°" not in df_pv_allow.columns:
        if "index" in df_pv_allow.columns:
            df_pv_allow = df_pv_allow.rename(columns={"index": "N°"})

    # Start writing on sheet: we'll keep track current row
    r = 2

    # --- TABLE 1 : Cost salary (Facture HT per employee per month) ---
    ws.cell(row=r, column=2, value="Cost salary - Invoice before VAT per employee").font = Font(bold=True, size=12)
    r += 1

    # header
    ws.cell(row=r, column=2, value="Employee").font = header_font
    ws.cell(row=r, column=2).fill = header_fill
    ws.cell(row=r, column=2).alignment = center
    for i, m in enumerate(mois_ordre, start=3):
        ws.cell(row=r, column=i, value=m).font = header_font
        ws.cell(row=r, column=i).fill = header_fill
        ws.cell(row=r, column=i).alignment = center

    r += 1
    start_data_row = r

    # lignes employés
    for eid in emp_ids:
        # nom
        nom = None
        for df_ in (df_piv, df_pv_travel, df_pv_allow):
            if "N°" in df_.columns and eid in df_["N°"].values:
                row0 = df_[df_["N°"] == eid].iloc[0]
                nom = row0.get("Nom", None) or row0.get("Nom_x", None) or str(eid)
                break
        if nom is None:
            nom = str(eid)

        ws.cell(row=r, column=2, value=str(nom)).alignment = Alignment(horizontal="center")


        monthly_vals = []
        total_ht = 0

        for c_idx, col in enumerate(facture_cols, start=3):
            val = df_piv.loc[df_piv["N°"] == eid, col].values[0] if (col in df_piv.columns and eid in df_piv["N°"].values) else 0.0
            val = float(val or 0)
            monthly_vals.append(val)
            ws.cell(row=r, column=c_idx, value=val).alignment = center
            total_ht += val

        # stocker pour calcul par ligne plus tard
        df_piv.loc[df_piv["N°"] == eid, "_TotalHT"] = total_ht
        df_piv.loc[df_piv["N°"] == eid, "_TVA"] = total_ht * (tva_pct/100)

        taux_for_row = None
        for i_m, v in enumerate(monthly_vals):
            if v != 0:
                month_name = mois_ordre[i_m]
                taux_for_row = taux_data.get(month_name, {}).get(devise)
                if taux_for_row:
                    break

        df_piv.loc[df_piv["N°"] == eid, "_Taux"] = taux_for_row or ""

        if taux_for_row:
            df_piv.loc[df_piv["N°"] == eid, "_HTdevise"] = total_ht / taux_for_row
            df_piv.loc[df_piv["N°"] == eid, "_TTCdevise"] = (total_ht / taux_for_row) * (1 + tva_pct/100)
        else:
            df_piv.loc[df_piv["N°"] == eid, "_HTdevise"] = None
            df_piv.loc[df_piv["N°"] == eid, "_TTCdevise"] = None

        r += 1

    end_data_row = r - 1
    r += 1

    # --- EXTRA LIGNES (au lieu d'extra colonnes !) ---
    def write_extra_line(label, colname):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = bold
        for i_m, m in enumerate(mois_ordre, start=3):
            # SUM uniquement sur la colonne des employés
            col_letter = get_column_letter(i_m)
            ws.cell(row=r, column=i_m,
                    value=f"=SUM({col_letter}{start_data_row}:{col_letter}{end_data_row})"
                    if colname is None else "")
        r += 1

    # 1) Total HT déjà calculé :
    ws.cell(row=r, column=2, value="Invoice before VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_red
    for i_m, m in enumerate(mois_ordre, start=3):
        cell_ht = ws.cell(row=r, column=i_m,
                    value=f"=SUM({get_column_letter(i_m)}{start_data_row}:{get_column_letter(i_m)}{end_data_row})")
        cell_ht.font = font_white_bold
        cell_ht.fill = fill_red 
        cell_ht.alignment = center
        
    salary_totals_row = r

    r += 1

    # 2) TVA = HT * 19%
    ws.cell(row=r, column=2, value="Invoice including VAT").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m, m in enumerate(mois_ordre, start=3):
        cell_ttc=ws.cell(row=r, column=i_m,
                value=f"={get_column_letter(i_m)}{r-1}+{get_column_letter(i_m)}{r-1}*{tva_pct/100}")
        cell_ttc.font = bold
        cell_ttc.fill = fill_sky_blue
        cell_ttc.alignment = center
    r += 1

    # 3) Taux change (ligne unique)
    ws.cell(row=r, column=2, value="Change Rate").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue

    for i_m, m in enumerate(mois_ordre, start=3):
        cell_rate=ws.cell(row=r, column=i_m, value=taux_data.get(m, {}).get(devise, ""))
        cell_rate.font = bold
        cell_rate.fill = fill_sky_blue
        cell_rate.alignment = center
    r += 1

    # 4) HT en devise
    ws.cell(row=r, column=2, value="USD equivalent before VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_dark_blue
    for i_m, m in enumerate(mois_ordre, start=3):
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            cell_usd_ht=ws.cell(row=r, column=i_m, value=f"={get_column_letter(i_m)}{r-3}/{taux}")
            cell_usd_ht.font = font_white_bold
            cell_usd_ht.fill = fill_dark_blue
            cell_usd_ht.alignment = center
    r += 1

    # 5) TTC en devise
    ws.cell(row=r, column=2, value="USD equivalent including VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_dark_blue
    for i_m, m in enumerate(mois_ordre, start=3):
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            cell_usd_ttc=ws.cell(row=r, column=i_m, value=f"={get_column_letter(i_m)}{r-3}/{taux}")
            cell_usd_ttc.font = font_white_bold
            cell_usd_ttc.fill = fill_dark_blue
            cell_usd_ttc.alignment = center
    r += 3


    # --- TABLE 2 : Travel expenses (M & C) ---
    ws.cell(row=r, column=2, value="Travel expenses (M & C)").font = Font(bold=True, size=12)
    r += 1

    # HEADER principal
    ws.cell(row=r, column=2, value="Employee").font = header_font
    ws.cell(row=r, column=2).fill = header_fill
    col = 3
    for m in mois_ordre:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        cell = ws.cell(row=r, column=col, value=m)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        col += 2

    r += 1
    start_travel_rows = r

    # lignes employés
    for eid in emp_ids:
        # Nom
        nom = df_piv[df_piv["N°"] == eid]["Nom"].values[0] if eid in df_piv["N°"].values else str(eid)
        ws.cell(row=r, column=2, value=nom).alignment = center

        col = 3
        sum_m = 0
        sum_c = 0

        for i_m, m in enumerate(mois_ordre):
            col_m = travel_m_cols[i_m]
            col_c = travel_c_cols[i_m]

            val_m = df_pv_travel.loc[df_pv_travel["N°"]==eid, col_m].values[0] if (col_m in df_pv_travel and eid in df_pv_travel["N°"].values) else 0
            val_c = df_pv_travel.loc[df_pv_travel["N°"]==eid, col_c].values[0] if (col_c in df_pv_travel and eid in df_pv_travel["N°"].values) else 0

            sum_m += float(val_m or 0)
            sum_c += float(val_c or 0)

            ws.cell(row=r, column=col, value=val_m).alignment = center
            ws.cell(row=r, column=col+1, value=val_c).alignment = center
            col += 2

        r += 1

    end_travel_rows = r - 1
    r += 1
    # --- EXTRA LIGNES TABLEAU 2 ---

    # Ligne Totaux Travel (M + C)
    ws.cell(row=r, column=2, value="Invoice before VAT  M/C segment").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_red

    col = 3   # première colonne M (Avril)

    for i_m in range(len(mois_ordre)):

        # --- TOTAL M pour ce mois ---
        col_letter_M = get_column_letter(col)
        cell_m_ht=ws.cell(
            row=r,
            column=col,
            value=f"=SUM({col_letter_M}{start_travel_rows}:{col_letter_M}{end_travel_rows})"
        )
        cell_m_ht.fill = fill_red
        cell_m_ht.font = font_white_bold
        cell_m_ht.alignment = center
        # --- TOTAL C pour ce mois ---
        col_letter_C = get_column_letter(col + 1)
        cell_c_ht=ws.cell(
            row=r,
            column=col + 1,
            value=f"=SUM({col_letter_C}{start_travel_rows}:{col_letter_C}{end_travel_rows})"
        )
        cell_c_ht.fill = fill_red
        cell_c_ht.font = font_white_bold
        cell_c_ht.alignment = center
        # passer au mois suivant (M puis C => +2 colonnes)
        col += 2
    
    r += 1

    # --- 3) Taxe 5% (séparée pour M et C) ---
    ws.cell(row=r, column=2, value=f"Taxe").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m in range(len(mois_ordre)):
        col_M = get_column_letter(3 + i_m*2)
        col_C = get_column_letter(4 + i_m*2)

        # Taxe sur M
        cell_m_tax=ws.cell(row=r, column=3 + i_m*2,
                value=f"={col_M}{r-1}+{col_M}{r-1}*{apply_special_tax_pct}/100")
        cell_m_tax.font= bold
        cell_m_tax.fill= fill_sky_blue
        cell_m_tax.alignment = center
        # Taxe sur C
        cell_c_tax=ws.cell(row=r, column=4 + i_m*2,
                value=f"={col_C}{r-1}+{col_C}{r-1}*{apply_special_tax_pct}/100")
        cell_c_tax.font= bold
        cell_c_tax.fill= fill_sky_blue
        cell_c_tax.alignment = center
    travel_totals_row = r
    r += 1


    # --- 4) TVA 19% (séparée pour M et C) ---
    ws.cell(row=r, column=2, value=f"Invoice including VAT").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m in range(len(mois_ordre)):
        col_M = get_column_letter(3 + i_m*2)
        col_C = get_column_letter(4 + i_m*2)

        # TVA sur M
        cell_m_tva=ws.cell(row=r, column=3 + i_m*2,
                value=f"={col_M}{r-1}+{col_M}{r-1}*{tva_pct}/100")
        cell_m_tva.font=bold
        cell_m_tva.fill= fill_sky_blue
        cell_m_tva.alignment = center

        # TVA sur C
        cell_c_tva=ws.cell(row=r, column=4 + i_m*2,
                value=f"={col_C}{r-1}+{col_C}{r-1}*{tva_pct}/100")
        cell_c_tva.font=bold
        cell_c_tva.fill= fill_sky_blue
        cell_c_tva.alignment = center

    r += 1


    # --- 5) Taux de change (une seule ligne mais répétée sur M et C) ---
    ws.cell(row=r, column=2, value="change rate").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise, "")
        cell_rate_m=ws.cell(row=r, column=3 + i_m*2, value=taux)
        cell_rate_c=ws.cell(row=r, column=4 + i_m*2, value=taux)  # pour C aussi
        cell_rate_m.font = bold
        cell_rate_m.fill = fill_sky_blue
        cell_rate_m.alignment = center
        cell_rate_c.font = bold
        cell_rate_c.fill = fill_sky_blue
        cell_rate_c.alignment = center

    r += 1


    # --- 6) HT en devise (séparé pour M et C) ---
    ws.cell(row=r, column=2, value="USD equivalent before VAT ").font = font_white_bold
    ws.cell(row=r, column=2).fill= fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            col_M = get_column_letter(3 + i_m*2)
            col_C = get_column_letter(4 + i_m*2)

            # M en devise
            cell_m_usd=ws.cell(row=r, column=3 + i_m*2,
                    value=f"={col_M}{r-3}/{taux}")
            cell_m_usd.font= font_white_bold
            cell_m_usd.fill= fill_dark_blue
            cell_m_usd.alignment = center

            # C en devise
            cell_c_usd=ws.cell(row=r, column=4 + i_m*2,
                    value=f"={col_C}{r-3}/{taux}")
            cell_c_usd.font= font_white_bold
            cell_c_usd.fill= fill_dark_blue
            cell_c_usd.alignment = center
    r += 1


    # --- 7) TTC en devise (séparé pour M et C) ---
    ws.cell(row=r, column=2, value="USD equevalent including VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill= fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            col_M = get_column_letter(3 + i_m*2)
            col_C = get_column_letter(4 + i_m*2)

            # M en devise
            cell_m_usd_ttc=ws.cell(row=r, column=3 + i_m*2,
                    value=f"={col_M}{r-4}*(1+{tva_pct}/100)")
            cell_m_usd_ttc.font= font_white_bold
            cell_m_usd_ttc.fill= fill_dark_blue
            cell_m_usd_ttc.alignment = center

            # C en devise
            cell_c_usd_ttc=ws.cell(row=r, column=4 + i_m*2,
                    value=f"={col_C}{r-4}*(1+{tva_pct}/100)")
            cell_c_usd_ttc.font= font_white_bold
            cell_c_usd_ttc.fill= fill_dark_blue
            cell_c_usd_ttc.alignment = center

    r += 3
    # ----------------------------------------------------
    # TABLEAU 3 — ALLOWANCE (M et C)
    # ----------------------------------------------------

    ws.cell(row=r, column=2, value="Allowance (M & C)").font = Font(bold=True, size=12)
    r += 1

    # HEADER principal
    ws.cell(row=r, column=2, value="Employee").font = header_font
    ws.cell(row=r, column=2).fill = header_fill
    col = 3
    for m in mois_ordre:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        cell = ws.cell(row=r, column=col, value=m)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        col += 2

    r += 1

    start_allow_rows = r

    # ---- Lignes employés ----
    # lignes employés
    for eid in emp_ids:
        # Nom
        nom = df_piv[df_piv["N°"] == eid]["Nom"].values[0] if eid in df_piv["N°"].values else str(eid)
        ws.cell(row=r, column=2, value=nom).alignment = center

        col = 3
        sum_m = 0
        sum_c = 0

        for i_m, m in enumerate(mois_ordre):
            col_m = allow_m_cols[i_m]
            col_c = allow_c_cols[i_m]

            val_m = df_pv_allow.loc[df_pv_allow["N°"]==eid, col_m].values[0] if (col_m in df_pv_allow and eid in df_pv_allow["N°"].values) else 0
            val_c = df_pv_allow.loc[df_pv_allow["N°"]==eid, col_c].values[0] if (col_c in df_pv_allow and eid in df_pv_allow["N°"].values) else 0

            sum_m += float(val_m or 0)
            sum_c += float(val_c or 0)

            ws.cell(row=r, column=col, value=val_m).alignment = center
            ws.cell(row=r, column=col+1, value=val_c).alignment = center
            col += 2

        r += 1

    end_allow_rows = r - 1
    r += 1

    # ---- TOTAL M & C sur la même ligne ----
    # Ligne Totaux Travel (M + C)
    ws.cell(row=r, column=2, value="Invoice before VAT  M and C segment").font = font_white_bold
    ws.cell(row=r, column=2).fill= fill_red
    col = 3   # première colonne M (Avril)

    for i_m in range(len(mois_ordre)):

        # --- TOTAL M pour ce mois ---
        col_letter_M = get_column_letter(col)
        cell_m_allow_ht=ws.cell(
            row=r,
            column=col,
            value=f"=SUM({col_letter_M}{start_allow_rows}:{col_letter_M}{end_allow_rows})"
        )
        cell_m_allow_ht.font = font_white_bold
        cell_m_allow_ht.fill = fill_red
        cell_m_allow_ht.alignment = center
        # --- TOTAL C pour ce mois ---
        col_letter_C = get_column_letter(col + 1)
        cell_c_allow_ht=ws.cell(
            row=r,
            column=col + 1,
            value=f"=SUM({col_letter_C}{start_allow_rows}:{col_letter_C}{end_allow_rows})"
        )
        cell_c_allow_ht.font = font_white_bold
        cell_c_allow_ht.fill = fill_red
        cell_c_allow_ht.alignment = center
        # passer au mois suivant (M puis C => +2 colonnes)
        col += 2
    
    r += 1
    # --- 3) Taxe 5% (séparée pour M et C) ---
    ws.cell(row=r, column=2, value=f"Taxe").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m in range(len(mois_ordre)):
        col_M = get_column_letter(3 + i_m*2)
        col_C = get_column_letter(4 + i_m*2)

        # Taxe sur M
        cell_m_allow_tax=ws.cell(row=r, column=3 + i_m*2,
                value=f"={col_M}{r-1}+{col_M}{r-1}*{apply_special_tax_pct}/100")
        cell_m_allow_tax.font=bold
        cell_m_allow_tax.fill=fill_sky_blue
        cell_m_allow_tax.alignment = center

        # Taxe sur C
        cell_c_allow_tax=ws.cell(row=r, column=4 + i_m*2,
                value=f"={col_C}{r-1}+{col_C}{r-1}*{apply_special_tax_pct}/100")
        cell_c_allow_tax.font=bold
        cell_c_allow_tax.fill=fill_sky_blue
        cell_c_allow_tax.alignment = center
    allow_totals_row = r
    r += 1


    # --- 4) TVA 19% (séparée pour M et C) ---
    ws.cell(row=r, column=2, value=f"Invoice including VAT").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m in range(len(mois_ordre)):
        col_M = get_column_letter(3 + i_m*2)
        col_C = get_column_letter(4 + i_m*2)

        # TVA sur M
        cell_m_allow_ttc=ws.cell(row=r, column=3 + i_m*2,
                value=f"={col_M}{r-1}+{col_M}{r-1}*{tva_pct}/100")
        cell_m_allow_ttc.font= bold
        cell_m_allow_ttc.fill= fill_sky_blue
        cell_m_allow_ttc.alignment = center

        # TVA sur C
        cell_c_allow_ttc=ws.cell(row=r, column=4 + i_m*2,
                value=f"={col_C}{r-1}+{col_C}{r-1}*{tva_pct}/100")
        cell_c_allow_ttc.font= bold
        cell_c_allow_ttc.fill= fill_sky_blue
        cell_c_allow_ttc.alignment = center

    r += 1


    # --- 5) Taux de change (une seule ligne mais répétée sur M et C) ---
    ws.cell(row=r, column=2, value="change rate").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise, "")
        cell_rate_m_alloz=ws.cell(row=r, column=3 + i_m*2, value=taux)
        cell_rate_c_alloz=ws.cell(row=r, column=4 + i_m*2, value=taux)  # pour C aussi
        cell_rate_m_alloz.font = bold
        cell_rate_m_alloz.fill = fill_sky_blue
        cell_rate_m_alloz.alignment = center
        cell_rate_c_alloz.font = bold
        cell_rate_c_alloz.fill = fill_sky_blue
        cell_rate_c_alloz.alignment = center
    r += 1


    # --- 6) HT en devise (séparé pour M et C) ---
    ws.cell(row=r, column=2, value="USD  equivalent before VAT ").font = font_white_bold
    ws.cell(row=r, column=2).fill= fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            col_M = get_column_letter(3 + i_m*2)
            col_C = get_column_letter(4 + i_m*2)

            # M en devise
            cell_m_allow_usd=ws.cell(row=r, column=3 + i_m*2,
                    value=f"={col_M}{r-3}/{taux}")
            cell_m_allow_usd.font= font_white_bold
            cell_m_allow_usd.fill= fill_dark_blue
            cell_m_allow_usd.alignment = center
            # C en devise
            cell_c_allow_usd=ws.cell(row=r, column=4 + i_m*2,
                    value=f"={col_C}{r-3}/{taux}")
            cell_c_allow_usd.font= font_white_bold
            cell_c_allow_usd.fill= fill_dark_blue
            cell_c_allow_usd.alignment = center

    r += 1


    # --- 7) TTC en devise (séparé pour M et C) ---
    ws.cell(row=r, column=2, value="USD equevalent including VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill= fill_dark_blue

    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            col_M = get_column_letter(3 + i_m*2)
            col_C = get_column_letter(4 + i_m*2)

            # M en devise
            cell_m_allow_usd_ttc=ws.cell(row=r, column=3 + i_m*2,
                    value=f"={col_M}{r-4}*(1+{tva_pct}/100)")
            cell_m_allow_usd_ttc.font= font_white_bold
            cell_m_allow_usd_ttc.fill= fill_dark_blue
            cell_m_allow_usd_ttc.alignment = center
            # C en devise
            cell_c_allow_usd_ttc=ws.cell(row=r, column=4 + i_m*2,
                    value=f"={col_C}{r-4}*(1+{tva_pct}/100)")
            cell_c_allow_usd_ttc.font= font_white_bold
            cell_c_allow_usd_ttc.alignment = center
            cell_c_allow_usd_ttc.fill= fill_dark_blue

    r += 3


    # --- TABLE 4 : Phone package 7 lignes ---
    ws.cell(row=r, column=2, value="Phone package summary").font = Font(bold=True, size=12)
    
    r += 1
    # header months
    ws.cell(row=r, column=2, value="Libellé").font = header_font
    for i, m in enumerate(mois_ordre, start=3):
        ws.cell(row=r, column=i, value=m).font = header_font
        ws.cell(row=r, column=i).fill = header_fill
        ws.cell(row=r, column=i).alignment = center
    r += 1

    # lines 1..7 as specified
    # 1: Gross Phone package charges DZD 24000 each month
    ws.cell(row=r, column=2, value="Gross Phone package charges (DZD)").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i in range(len(mois_ordre)):
        gross_phone_cell=ws.cell(row=r, column=3+i, value=phone_gross_dzd)
        gross_phone_cell.font = bold
        gross_phone_cell.fill= fill_sky_blue
        gross_phone_cell.alignment = center
    r += 1

    # 2: taxe %
    ws.cell(row=r, column=2, value="Taxe").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i in range(len(mois_ordre)):
        cell_tax=ws.cell(row=r, column=3+i, value=(tax_pct_line2 if tax_pct_line2 is not None else ""))
        cell_tax.font= bold
        cell_tax.fill = fill_sky_blue
        cell_tax.alignment = center
    r += 1

    # 3: apply special tax (apply_special_tax_pct)
    ws.cell(row=r, column=2, value=f"Invoice before VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill= fill_red
    for i in range(len(mois_ordre)):
        phone_ht=ws.cell(row=r, column=3+i, value=phone_gross_dzd+(phone_gross_dzd * (apply_special_tax_pct/100)))
        phone_ht.font= font_white_bold
        phone_ht.fill = fill_red
        phone_ht.alignment = center
    phone_gross_row = r
    r += 1

    # 4: TVA 19%
    ws.cell(row=r, column=2, value=f"Invoice including VAT").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i in range(len(mois_ordre)):
        col_letter = get_column_letter(3 + i)
        # TVA = valeur de la ligne précédente * taux
        phone_ttc=ws.cell(
            row=r,
            column=3+i,
            value=f"={col_letter}{r-1}+{col_letter}{r-1}*{tva_pct/100}"
        )
        phone_ttc.font= bold
        phone_ttc.fill = fill_sky_blue
        phone_ttc.alignment = center
    r += 1


    # 5: taux de change (per month)
    ws.cell(row=r, column=2, value="Change rate").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m, m in enumerate(mois_ordre):
        taux_for_month = taux_data.get(m, {}).get(devise)
        cell_rate2=ws.cell(row=r, column=3+i_m, value=taux_for_month or "")
        cell_rate2.font = bold
        cell_rate2.fill = fill_sky_blue
        cell_rate2.alignment = center
    r += 1

    # 6: cout HT en devise
    ws.cell(row=r, column=2, value="USD equivalent before VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        taux_for_month = taux_data.get(m, {}).get(devise)
        col_letter = get_column_letter(3 + i_m)

        if taux_for_month:
            cell_ht_usd_phone=ws.cell(row=r, column=3+i_m,
                    value=f"={col_letter}{r-5}/{taux_for_month}")
            cell_ht_usd_phone.font = font_white_bold
            cell_ht_usd_phone.fill = fill_dark_blue
            cell_ht_usd_phone.alignment = center
        else:
            cell_ht_usd_phone=ws.cell(row=r, column=3+i_m, value=None)
            cell_ht_usd_phone.font = font_white_bold
            cell_ht_usd_phone.fill = fill_dark_blue
            cell_ht_usd_phone.alignment = center
    r += 1


    # 7: cout TTC en devise (HT * (1+TVA))
    ws.cell(row=r, column=2, value="USD equivalent including VAT").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        taux_for_month = taux_data.get(m, {}).get(devise)
        col_letter = get_column_letter(3 + i_m)
        if taux_for_month:
            cell_ttc_usd_phone=ws.cell(row=r, column=3+i_m,
                    value=f"={col_letter}{r-3}/{taux_for_month}")
            cell_ttc_usd_phone.font = font_white_bold
            cell_ttc_usd_phone.fill= fill_dark_blue
            cell_ttc_usd_phone.alignment = center
        else:
            cell_ttc_usd_phone=ws.cell(row=r, column=3+i_m, value=None)
            cell_ttc_usd_phone.font = font_white_bold
            cell_ttc_usd_phone.fill= fill_dark_blue
            cell_ttc_usd_phone.alignment = center
    r += 1
    r += 3
    # wide columns
    # ---------------------------------------------------------------
    #  TABLEAU FINAL GLOBAL
    # ---------------------------------------------------------------
    ws.cell(row=r, column=2, value="GLOBAL SUMMARY").font = Font(bold=True, size=12)
    r += 1
    # -----------------------------
    # 1) TOTAL GLOBAL PAR MOIS
    # -----------------------------
    ws.cell(row=r, column=2, value="Invoice before VAT ").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_red
    for i_m, m in enumerate(mois_ordre):
        # colonnes pour ce mois :
        # - Salary/Cost table uses single column per month: col_salary = 3 + i_m
        # - Travel & Allowance use 2 cols per month: M at 3 + i_m*2, C at 4 + i_m*2
        prev = max(i_m - 1, 0)
        col_salary = 3 + i_m
        col_allow_M  = 3 + prev * 2
        col_allow_C  = 4 + prev * 2
        col_travel_M = 3 + prev * 2
        col_travel_C = 4 + prev * 2
        col_phone = 3 + i_m

        # build formula that sums M and C for allow + travel + salary + phone
        formula = (
            "="
            f"{get_column_letter(col_allow_M)}{allow_totals_row}+"
            f"{get_column_letter(col_allow_C)}{allow_totals_row}+"
            f"{get_column_letter(col_travel_M)}{travel_totals_row}+"
            f"{get_column_letter(col_travel_C)}{travel_totals_row}+"
            f"{get_column_letter(col_salary)}{salary_totals_row}+"
            f"{get_column_letter(col_phone)}{phone_gross_row}"
        )

        ws.cell(row=r, column=3 + i_m, value=formula)
        ws.cell(row=r, column=3 + i_m).fill= fill_red
        ws.cell(row=r, column=3 + i_m).font= font_white_bold
        ws.cell(row=r, column=3 + i_m).alignment = center
    r += 1

    # -----------------------------
    # 2) VAT 19% (appliquée sur la ligne Total global)
    # -----------------------------
    ws.cell(row=r, column=2, value=f"Total Invoice including VAT (USD)").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m, m in enumerate(mois_ordre):
        col = 3 + i_m
        # la ligne précédente (total global) est r-1
        global_ttc=ws.cell(row=r, column=col, value=f"={get_column_letter(col)}{r-1}*{tva_pct/100}")
        global_ttc.font = bold
        global_ttc.fill = fill_sky_blue
        global_ttc.alignment = center
    r += 1

    # -----------------------------
    # 3) TAUX DE CHANGE
    # -----------------------------
    ws.cell(row=r, column=2, value="change rate").font = bold
    ws.cell(row=r, column=2).fill = fill_sky_blue
    for i_m, m in enumerate(mois_ordre):
        taux = taux_data.get(m, {}).get(devise, "")
        cell_rate3=ws.cell(row=r, column=3 + i_m, value=taux)
        cell_rate3.font = bold
        cell_rate3.fill = fill_sky_blue
        cell_rate3.alignment = center
    r += 1

    # -----------------------------
    # 4) HT EN DEVISE
    # -----------------------------
    ws.cell(row=r, column=2, value="Total Invoice before VAT (USD)").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        col = 3 + i_m
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            # HT en devise = (Total global in local currency) / taux
            cell_ht_usd_global=ws.cell(row=r, column=col, value=f"={get_column_letter(col)}{r-3}/{taux}")
            cell_ht_usd_global.font = font_white_bold
            cell_ht_usd_global.fill = fill_dark_blue
            cell_ht_usd_global.alignment = center
    r += 1

    # -----------------------------
    # 5) TTC EN DEVISE
    # -----------------------------
    ws.cell(row=r, column=2, value="Total Invoice Including VAT (USD)").font = font_white_bold
    ws.cell(row=r, column=2).fill = fill_dark_blue
    for i_m, m in enumerate(mois_ordre):
        col = 3 + i_m
        taux = taux_data.get(m, {}).get(devise)
        if taux:
            # TTC en devise = HT_en_devise * (1 + TVA)
            # HT_en_devise is in row r-1 (just written above), so formula uses that cell
            cell_ttc_usd_global=ws.cell(row=r, column=col, value=f"={get_column_letter(col)}{r-3}*(1+{tva_pct}/100)")
            cell_ttc_usd_global.font= font_white_bold
            cell_ttc_usd_global.fill = fill_dark_blue
            cell_ttc_usd_global.alignment = center
    r += 2




    # freeze top rows
    # ws.freeze_panes = "C3"

    return ws



def generer_facture_excel_sheet(employe_dict, ws, wb, logos_folder="Logos"):
     # 📌 Styles
    header_font = Font(bold=True, size=14, color="000000")
    normal_font_black = Font(size=11, color="000000")
    normal_font_white = Font(size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    COL_OFFSET = 4
    
    # 📌 Mapping couleurs
    color_map = {
        "Base cotisable": "9fc5e8", "Retenue CNAS employé": "9fc5e8",
        "Base imposable au baréme": "9fc5e8", "IRG barème": "9fc5e8",
        "Base imposable 10%": "9fc5e8", "IRG 10%": "9fc5e8",
        "Salaire brut": "9fc5e8", "CNAS employeur": "9fc5e8",
        "Cotisation œuvre sociale": "9fc5e8", "Taxe formation": "9fc5e8",
        "Taxe formation et os": "9fc5e8", 
        "Coût congé payé": "9fc5e8", "Taux complément santé (DZD)": "9fc5e8",
        "Fees etalent": "9fc5e8", "TAP": "9fc5e8",
        "Salaire net": "25488e", "Masse salariale": "25488e", "Coût salaire": "25488e",
        "Facture HT": "e11b36","Facture TVA": "284052", "Facture TTC": "284052",
    }
    white_text_lines = {"Salaire net", "Masse salariale", "Coût salaire",
                        "Facture HT", "Facture TVA", "Facture TTC"}
    

    # ---------------------- LOGO ----------------------
    etablissement = str(employe_dict.get("Etablissement", "")).strip()
    logo_path = os.path.join(logos_folder, f"{etablissement}.png")

    if os.path.exists(logo_path):
        try:
            logo = XLImage(logo_path)
            logo.width = 350
            logo.height = 120
            ws.add_image(logo, f"{get_column_letter(COL_OFFSET+4)}1")
        except:
            pass

    # ---------------------- INFOS EMPLOYE ----------------------
    infos_employe = [
        ["Nom:", employe_dict.get("Nom", "")],
        ["Prénom:", employe_dict.get("Prénom", "")],
        ["Année:", employe_dict.get("Année", "")],
        ["Titre du poste:", employe_dict.get("Titre du poste", "")],
        ["Durée CDD:", employe_dict.get("Durée du CDD (Mois)", "")],
        ["Établissement:", etablissement]
    ]

    for i, (label, value) in enumerate(infos_employe, start=3):
        ws.cell(row=i, column=COL_OFFSET, value=label).font = Font(bold=True)
        ws.cell(row=i, column=COL_OFFSET+1, value=value).font = normal_font_black
    # 📌 Catégorisation clients
        
        client_gd = ["G+D"]
    # ---------------------- EXTRACTION DES MOIS ----------------------
    mois_data = {}
    MOIS_VALIDES = {
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    }

    for key, value in employe_dict.items():
        if "_" in key:
            ligne_nom, mois_nom = key.rsplit("_", 1)
            if mois_nom not in MOIS_VALIDES:  
                continue   # ignore x, norm, ou autres colonnes parasites
            mois_data.setdefault(mois_nom, {})[ligne_nom] = value

    def is_useful_value(v):
        if v is None:
            return False
        try:
            if isinstance(v, float) and math.isnan(v):
                return False
        except:
            pass
        if isinstance(v, str):
            if v.strip().lower() in ("", "nan", "none", "null"):
                return False
            return True
        return True

    mois_disponibles = [m for m, lignes in mois_data.items()
                        if any(is_useful_value(v) for v in lignes.values())]

    # ---------------------- TABLEAU ----------------------
    def generer_tableau(start_row, titre, lignes):
        # Titre tableau
        ws.merge_cells(start_row=start_row, start_column=COL_OFFSET, 
                       end_row=start_row, end_column=COL_OFFSET+len(mois_disponibles))
        titre_cell = ws.cell(row=start_row, column=COL_OFFSET, value=titre)
        titre_cell.font = Font(bold=True, size=13, color="000000")
        titre_cell.alignment = center_alignment
        start_row += 1

        # En-tête
        ws.cell(row=start_row, column=COL_OFFSET, value="Éléments").font = header_font
        ws.cell(row=start_row, column=COL_OFFSET).fill = header_fill
        ws.cell(row=start_row, column=COL_OFFSET).alignment = center_alignment
        ws.cell(row=start_row, column=COL_OFFSET).border = border

        for col, mois in enumerate(mois_disponibles, start=1):
            cell = ws.cell(row=start_row, column=COL_OFFSET+col, value=mois)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Lignes
        for row, ligne in enumerate(lignes, start=1):
            current_row = start_row + row
            fill_color = color_map.get(ligne)
            fill_style = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
            font_color = "FFFFFF" if ligne in white_text_lines else "000000"

            cell_titre = ws.cell(row=current_row, column=COL_OFFSET, value=ligne)
            cell_titre.font = Font(bold=True, color=font_color)
            cell_titre.alignment = left_alignment
            cell_titre.border = border
            if fill_style: cell_titre.fill = fill_style

            for col, mois in enumerate(mois_disponibles, start=1):
                val = mois_data.get(mois, {}).get(ligne, " ")
                if isinstance(val, (int, float)):
                    val = f"{val:,.2f}".replace(",", " ").replace(".", ",")
                cell = ws.cell(row=current_row, column=COL_OFFSET+col, value=val)
                cell.font = Font(size=11, color=font_color)
                cell.alignment = center_alignment
                cell.border = border
                if fill_style: cell.fill = fill_style
                elif row % 2 == 0: cell.fill = data_fill

        return current_row + 2  # ligne suivante

    # 📌 Cas spécial G+D → 3 tableaux
    
    start_row = 10
    start_row = generer_tableau(start_row, "Récapitulatif salarial",
            ["Salaire de base","Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Prime mensuelle", "Frais de remboursement (Véhicule) (DZD)",
             "Salaire net","Coût salaire", "Facture HT","Facture TVA", "Facture TTC","Taux de change","Facture HT en devise","Facture TTC en devise"])
    start_row = generer_tableau(start_row, "Travel Expenses",
            ["Travel expenses M segment", "Travel expenses C segment","Total Travel expenses","Taxe", "Total Travel expenses before VAT","Total Travel expenses including VAT","Taux de change", "Total Travel expenses en devise" ])
    start_row = generer_tableau(start_row, "Allowance",
        ["Allowance M segment", "Allowance C segment", "Total Allowance","Taxe","Total Allowance before VAT","Total Allowance Including VAT","Taux de change","Total Allowance en devise"])


    for col in range(COL_OFFSET, COL_OFFSET + len(mois_disponibles) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 35

    ws.freeze_panes = "E1"

    
def calcul_cout_conge(row):
    cout_conge = row["Coût congé payé"]
    return cout_conge
USERS_FILE = "users.json"
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
# 🔒 Hachage du mot de passe
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# 📝 Inscription
def signup(email, password):
    response = supabase.auth.sign_up({"email": email, "password": password})
    return response

# 🔐 Connexion
def login(email, password):
    response = supabase.auth.sign_in_with_password({"email": email, "password": password})
    return response

# 🚪 Déconnexion
def logout():
    supabase.auth.sign_out()
    st.session_state["authenticated"] = False
    st.session_state["user"] = None

# --- Interface ---
# st.title("🔐 Auth avec Supabase")

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    mode = st.radio("Choisissez :", ["Connexion", "Créer un compte"])

    email = st.text_input("Email")
    password = st.text_input("Mot de passe", type="password")

    if mode == "Créer un compte":
        if st.button("Créer un compte"):
            res = signup(email, password)
            if res.user:
                st.success("✅ Compte créé ! Vérifie ton email.")
            else:
                st.error(f"Erreur : {res}")
    else:
        if st.button("Connexion"):
            res = login(email, password)
            if res.user:
                st.session_state["authenticated"] = True
                st.session_state["user"] = res.user
                st.rerun()
            else:
                st.error("❌ Email ou mot de passe incorrect.")

else:
    with st.sidebar:
        st.success(f"Bienvenue {st.session_state['user'].email} 👋")
        if st.button("Se déconnecter"):
            logout()
            st.rerun()

# 🎉 Application principale ici
# st.title("Bienvenue sur l'application de calcule des factures")



    # 📌 Initialisation
    if "clients" not in st.session_state:
        st.session_state.clients = []
    if "selected_client" not in st.session_state:
        st.session_state.selected_client = None
    if "full_df" not in st.session_state:
        st.session_state.full_df = None
    if "data" not in st.session_state:
        st.session_state.data = {}


    CLIENTS_FILE = Path("clients.json")
    # st.title("👥 Gestion des clients et des employés")

    # 📂 Charger la liste des clients
    if CLIENTS_FILE.exists():
        with open(CLIENTS_FILE, "r", encoding="utf-8") as f:
            clients_list = json.load(f)
    else:
        clients_list = ["G+D"]
        with open(CLIENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(clients_list, f, ensure_ascii=False, indent=2)
    client_name = st.sidebar.selectbox(
        "Sélectionner un client",
        options=clients_list,
        index=None,  # <-- pas de sélection initiale
        placeholder="— Sélectionner un client —",
        key="client_select",
    )
    st.session_state.clients = clients_list
    st.session_state.selected_client = client_name
    # 📁 Upload du fichier global
    # GLOBAL_CSV = "data_global.csv"
    st.sidebar.subheader("📅 Charger le fichier Récap")
    uploaded_csv = st.sidebar.file_uploader("Fichier CSV Récap", type=["csv"], key="csv_recap")
    
    MOIS_MAP = {
    "-janv.-": "Janvier",
    "-févr.-": "Février",
    "-mars-": "Mars",
    "-avr.-": "Avril",
    "-mai-": "Mai",
    "-juin-": "Juin",
    "-juil.-": "Juillet",
    "-août-": "Août",
    "-sept.-": "Septembre",
    "-oct.-": "Octobre",
    "-nov.-": "Novembre",
    "-déc.-": "Décembre",
    }

    if uploaded_csv is not None:
        try:
            raw = uploaded_csv.getvalue()

            # Corriger les espaces insécables (U+202F)
            raw = raw.replace(b"\xe2\x80\xaf", b" ")

            # Charger CSV
            df_full = pd.read_csv(
                io.BytesIO(raw),
                skiprows=2,
                sep=",",
                decimal=",",
                thousands=" "
            )
            
            st.write(df_full.head())
            st.session_state.full_df = df_full
            st.sidebar.success("✅ Fichier chargé avec succès !")

            # 🚑 Nettoyer les NaN et inf (pour JSON)
            df_full = df_full.replace([np.nan, np.inf, -np.inf], None)

            # Vérifier et normaliser la colonne Mois
            if "Mois" in df_full.columns:
                df_full["Mois"] = df_full["Mois"].map(MOIS_MAP).fillna(df_full["Mois"])
                mois = df_full["Mois"].iloc[0]
            else:
                st.error("❌ La colonne 'Mois' est manquante dans le CSV")
                mois = "Inconnu"

            # Supprimer les anciennes lignes du même mois
            
            supabase.table("Recap").delete().eq("Mois", mois).execute()

            # 🔎 Récupérer le max(id) existant
            res = supabase.table("Recap").select("id").order("id", desc=True).limit(1).execute()
            start_id = res.data[0]["id"] + 1 if res.data else 1

            # Assigner des id uniques
            df_full.insert(0, "id", range(start_id, start_id + len(df_full)))

            # Insérer les nouvelles lignes par batch
            records = df_full.to_dict(orient="records")
            CHUNK = 1000
            for i in range(0, len(records), CHUNK):
                batch = records[i:i+CHUNK]
                if batch:
                    supabase.table("Recap").insert(batch).execute()

            st.success(f"✅ Données mises à jour pour **{mois}** dans Supabase (table Recap)")
            # st.dataframe(df_full)

        except Exception as e:
            st.sidebar.error(f"❌ Erreur : {e}")

    if st.session_state.selected_client:
        st.markdown(f"## 👤 Données des employés pour **{st.session_state.selected_client.strip()}**")

        if st.session_state.full_df is not None:
            # df = st.session_state.full_df.copy()
            

            df = fetch_all_data("Recap")
            # st.write("Nb lignes récupérées :", len(df))
            # st.write("Mois distincts :", sorted(df["Mois"].unique()))

            # ✅ Filtrer par client
            df_client, info = trouver_client_robuste(st.session_state.selected_client, df, debug=True)
            # st.write("🔍 Debug trouver_client_robuste:", info)
            # st.write("✅ Nombre de lignes trouvées pour le client:", len(df_client))

        # ⚠️ Ne pas écraser totalement, stocker tel quel


            if not df_client.empty:
                # ------------------------------------------------
                # Partie calculs et préparation des données
                # ------------------------------------------------
                mois_possibles = [mois.lower() for mois in calendar.month_name if mois]
                colonnes_mois = [col for col in df_client.columns if any(mois in col.lower() for mois in mois_possibles)]
                nb_employes = df_client["N°"].nunique()
                st.success(f"{nb_employes} employés trouvés.")
                source = st.sidebar.radio("📌 Source des taux :", ["Capture d'écran", "OANDA"])
                if "df_rates" not in st.session_state:
                    st.session_state.df_rates = pd.DataFrame(columns=["Devise", "Achat", "Vente"])
                    devise_active = st.sidebar.radio("Choisir la devise :", ["EUR", "USD"], horizontal=True)
                    st.session_state.devise_active = devise_active
                if source == "Capture d'écran":
                    from PIL import Image  
                    MAX_WIDTH = 800
                    uploaded_image = st.sidebar.file_uploader("📷 Charger la capture", type=["png", "jpg", "jpeg"])

                    if uploaded_image is not None:
                        image = Image.open(uploaded_image)

                        # Redimensionner si trop large
                        width, height = image.size
                        if width > MAX_WIDTH:
                            ratio = MAX_WIDTH / width
                            new_size = (MAX_WIDTH, int(height * ratio))
                            image = image.resize(new_size, Image.Resampling.LANCZOS)

                        st.sidebar.image(image, caption="Capture chargée", use_container_width=True)

                        # OCR avec EasyOCR
                        reader = easyocr.Reader(['en', 'fr'])
                        result = reader.readtext(np.array(image), detail=0)  # texte brut
                        lines = [line.strip() for line in result if line.strip()]

                        # st.sidebar.text_area("📄 Texte brut OCR", value="\n".join(lines), height=200)

                        # 🔎 Extraire les blocs Devise + Achat + Vente
                        data = []
                        i = 0
                        while i < len(lines):
                            line = lines[i]

                            # Devise = 3 lettres majuscules
                            if re.match(r"^[A-Z]{3}$", line):
                                devise = line
                                try:
                                    achat = float(lines[i+1].replace(",", "."))
                                    vente = float(lines[i+2].replace(",", "."))
                                    data.append([devise, achat, vente])
                                    i += 3  # avancer de 3 lignes
                                    continue
                                except Exception:
                                    pass
                            i += 1

                        if data:
                            df_rates = pd.DataFrame(data, columns=["Devise", "Achat", "Vente"])
                            st.write("📊 Taux extraits :")
                            st.dataframe(df_rates)

                            # Sauvegarde en session_state
                            st.session_state.df_rates = df_rates

                            # Interface choix devise
                            devise_active = st.sidebar.radio("Choisir la devise :", ["EUR", "USD"], horizontal=True)
                            st.session_state.devise_active = devise_active


                        else:
                            st.warning("⚠️ Aucun taux détecté dans l'image")
                elif source == "OANDA":
                    if "df_rates" not in st.session_state:
                        st.session_state.df_rates = pd.DataFrame(
                            [["EUR", 1.0, 1.0]],  # Valeur par défaut
                            columns=["Devise", "Achat", "Vente"]
                        )
                    st.sidebar.markdown("## 💱 Saisie du taux de change (OANDA manuel)")
                    with st.sidebar.form("oanda_manual_form"):
                        base = st.text_input("Devise de base", "USD").upper().strip()
                        target = st.text_input("Devise cible", "DZD").upper().strip()
                        taux = st.number_input(
                            f"Taux (1 {base} = ? {target})",
                            min_value=0.0,
                            format="%.6f"
                        )
                        submit = st.form_submit_button("Enregistrer")
                    devise_active = st.sidebar.radio("Choisir la devise :", ["USD", "EUR"], horizontal=True)
                    st.session_state.devise_active = devise_active
                    if submit:
                        if taux <= 0:
                            st.error("Le taux doit être supérieur à 0.")
                        else:
                            pair = f"{base}"
                            new_row = {"Devise": pair, "Achat": taux, "Vente": taux}

                            df = st.session_state.df_rates.copy()
                            mask = df["Devise"] == pair
                            if mask.any():
                                df.loc[mask, ["Achat", "Vente"]] = [taux, taux]
                            else:
                                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                            st.session_state.df_rates = df
                            st.sidebar.success(f"Taux {pair} enregistré : {taux}")

                col1, = st.columns(1) 
                with col1:
                    jours_mois = st.number_input("Jours mois", min_value=28.0, max_value=31.0, step=1.0, value=30.0)

                Observation = st.text_input("Observation", value="")

                # 2. Nettoyage des colonnes
                cols_to_float = [
                    "Salaire de base (DZD)", "Prime mensuelle (Barème) (DZD)", "IFSP (20% du salaire de base)",
                    "Prime exeptionnelle (10%) (DZD)", "Frais de remboursement (Véhicule) (DZD)", 
                    "Indemnité de panier (DZD)", "Indemnité de transport (DZD)", "Nouveau Salaire de base (DZD)",
                    "Prime vestimentaire (DZD)", "Nouvelle Indemnité de panier (DZD)",  "Nouvelle Indemnité de transport (DZD)",
                    "Nouvelle Prime mensuelle (DZD)", "Nouveaux Frais de remboursement (Véhicule) (DZD)","Prime vestimentaire (DZD)", "Indémnité Véhicule (DZD)",
                    "Absence (Jour)","Absence Maladie (Jour)","Absence Maternité (Jour)", "Absence Mise à pied (Jour)", "Jours de congé (Jour)",
                    "Heures supp 100% (H)", "Heures supp 75% (H)", "Heures supp 50% (H)", "Jours supp (Jour)","Taux complément santé (DZD)",
                   "Avance NET (DZD)", "Coût congé payé", "Nbr jours STC (jours)",
                    "Jours de congé ouvré(22 jours)","Indemnité non cotisable et imposable 10% (DZD)", "Total absence (sur 22 jours)",
                    "Nouvelle indémnité Véhicule (DZD)","Nouveau IFSP (20% du salaire de base)","Indemnité de départ (Net)",
                    "Allocation Aid El Adha NET"
                ]
              

                


                for col in cols_to_float:
                    if col in df_client.columns:
                        df_client[col] = nettoyer_colonne(df_client, col)
                    else:
                        df_client[col] = 0.0
                col_pourcentage = ["Fees etalent"]

                for col in col_pourcentage:
                    df_client[col] = (
                        df_client[col]
                        .fillna("0")                          
                        .astype(str)
                        .str.replace("%", "", regex=False)
                        .str.replace(",", ".", regex=False)
                        .str.strip()
                        .replace("", "0")
                        .astype(float)
                    )
                
             
                absences_total = (
                df_client["Absence (Jour)"]
                + df_client["Absence Maladie (Jour)"]
                + df_client["Absence Maternité (Jour)"]
                + df_client["Absence Mise à pied (Jour)"]
                + df_client["Jours de congé (Jour)"]  # version brute
                )
            
                absences_total22 = (
                    df_client["Total absence (sur 22 jours)"]
                    + df_client["Jours de congé ouvré(22 jours)"]  
                )
                
                
                HEURES_MOIS = 173.33
                
                # 3. Calculs (une seule fois)
                df_client["Salaire de base calcule"] = (get_valeur("Salaire de base (DZD)", "Nouveau Salaire de base (DZD)"))
                df_client["Indemnité de panier calcule"] = get_valeur("Indemnité de panier (DZD)", "Nouvelle Indemnité de panier (DZD)")
                df_client["indémnité Véhicule calcule"] = get_valeur("Indémnité Véhicule (DZD)", "Nouvelle indémnité Véhicule (DZD)")
                df_client["Indemnité de transport calcule"] = get_valeur("Indemnité de transport (DZD)", "Nouvelle Indemnité de transport (DZD)")
                df_client["Prime mensuelle calcule"] = get_valeur("Prime mensuelle (DZD)", "Nouvelle Prime mensuelle (DZD)")
                df_client["IFSP (20% du salaire de base) calcule"] = get_valeur("IFSP (20% du salaire de base)", "Nouveau IFSP (20% du salaire de base)")
                df_client["Frais remboursement calcule"] = get_valeur("Frais de remboursement (Véhicule) (DZD)", "Nouveaux Frais de remboursement (Véhicule) (DZD)")
                df_client["Salaire de base calcule"] += df_client["IFSP (20% du salaire de base) calcule"]
                salaire_journalier = df_client["Salaire de base calcule"] / jours_mois
                
                df_client["IFSP (20% du salaire de base)"] = df_client["IFSP (20% du salaire de base) calcule"]
                

                
                
                
                df_client["Indemnitésomme"]= df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] + df_client["Prime vestimentaire (DZD)"] + df_client["indémnité Véhicule calcule"]
                    # st.write(df_client["Indemnitésomme"].head(133))
                df_client["Indemnité 22jours"] = (
                        df_client["Indemnitésomme"]
                        - (df_client["Indemnitésomme"] / 22 * 0)
                        + (df_client["Indemnitésomme"] / 22 * (
                            (df_client["Heures supp 100% (H)"] * 0) / 8
                            + (df_client["Heures supp 75% (H)"] * 0) / 8
                            + (df_client["Heures supp 50% (H)"] * 0) / 8
                        ))
                    )
                   
                
                
                df_client["Base cotisable"] = (
                            
                         df_client["Salaire de base calcule"] + df_client["Prime mensuelle (Barème) (DZD)"]  + df_client["Indemnité non cotisable et imposable 10% (DZD)"]
                        )
                df_client["indémnité Véhicule"] = df_client["indémnité Véhicule calcule"]
                
                df_client["Base imposable 10%"] = df_client["Prime exeptionnelle (10%) (DZD)"] * 0.91

                df_client["Retenue CNAS employé"] = df_client["Base cotisable"] * 0.09
                df_client["Base imposable au baréme"] = np.floor((((df_client["Salaire de base calcule"] + df_client["Prime exeptionnelle (10%) (DZD)"]) -df_client["Prime exeptionnelle (10%) (DZD)"]) * 0.91 + df_client["Indemnité 22jours"])/10)*10
                  
                def irg_bareme(base):
                    b = np.ceil(base / 10) * 10  # PLAFOND(...;10) en Excel
                    
                    if b < 30000:
                        return 0
                    elif 30000 <= b <= 30860:
                        return (((( (b - 20000) * 0.23) - 1000) * 137/51) - (27925/8))
                    elif 30860 < b < 35000:
                        return (((( (b - 20000) * 0.23) * 0.60) * 137/51) - (27925/8))
                    elif 35000 <= b < 36310:
                        return (b - 20000) * 0.23 * 0.60
                    elif 36310 <= b < 40000:
                        return (b - 20000) * 0.23 - 1500
                    elif 40000 <= b < 80000:
                        return (b - 40000) * 0.27 + 3100
                    elif 80000 <= b < 160000:
                        return (b - 80000) * 0.30 + 13900
                    elif 160000 <= b < 320000:
                        return (b - 160000) * 0.33 + 37900
                    else:  # b >= 320000
                        return (b - 320000) * 0.35 + 90700

                df_client["IRG barème"] = df_client["Base imposable au baréme"].apply(irg_bareme)
                df_client["IRG 10%"] = df_client["Base imposable 10%"] * 0.10
                
                df_client["Salaire brut"] = (
                        (df_client["Base cotisable"] +
                        (df_client["Indemnité 22jours"])+
                        df_client["Frais remboursement calcule"]) 
                    )
                df_client["Salaire net"] = ((df_client["Salaire de base calcule"]+df_client["Indemnité 22jours"]+df_client["Prime exeptionnelle (10%) (DZD)"])-((df_client["Prime exeptionnelle (10%) (DZD)"]+df_client["Salaire de base calcule"])*0.09)-df_client["IRG barème"])
                
                df_client["CNAS employeur"] = df_client["Base cotisable"] * 0.26
                df_client["Indemnités Non Cotisable - Mensuelle | Panier, Transport"] = df_client["Indemnité 22jours"]
                
                df_client["Cotisation œuvre sociale"] = df_client["Salaire brut"] * 0.02
                df_client["Taxe formation"] = df_client["Salaire brut"] * 0.02
                df_client["Taxe formation et os"] = 0
                df_client["Masse salariale"] = (
                    df_client["Salaire brut"] +
                    df_client["CNAS employeur"] +
                    df_client["Cotisation œuvre sociale"] +
                    df_client["Taxe formation"]
                )
                df_client["Prime mensuelle"] = df_client["Prime mensuelle calcule"]
                
                        
                
                df_client["Coût salaire"] = (df_client["Salaire de base calcule"] + df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] +df_client["Prime vestimentaire (DZD)"]+df_client["Frais remboursement calcule"]+df_client["Prime exeptionnelle (10%) (DZD)"]) 
                    # df_client["Coût salaire"] += np.where(
                    # df_client["Base de régul"] == "Cout salaire", df_client["Régul"], 0)
                fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                if df_client["TAP"].iloc[0] == "Oui" :
                        df_client["TAP (DZD)"] = (df_client["Coût salaire"] + ( df_client["Coût salaire"] * df_client["Fees etalent"])) * 0.02
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur) + df_client["TAP (DZD)"])
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            import json
                            from pathlib import Path

                            # Chemin du JSON
                            json_path = Path("taux_change.json")

                            # Charger JSON existant ou créer un dictionnaire vide
                            if json_path.exists():
                                with open(json_path, "r", encoding="utf-8") as f:
                                    taux_data = json.load(f)
                            else:
                                taux_data = {}

                            # Exemple : df_rates contient la capture de change
                            # On récupère le mois actuel du système
                            from datetime import datetime
                            mois_courant_num = datetime.now().month
                            mois_list = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                                        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                            mois_courant = mois_list[mois_courant_num - 1]

                            # Récupérer les taux de la capture
                            eur_rate = df_rates.loc[df_rates["Devise"].str.contains("EUR"), "Achat"].values[0]
                            usd_rate = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Mettre à jour le JSON
                            taux_data[mois_courant] = {"EUR": float(eur_rate), "USD": float(usd_rate)}

                            # Écrire le JSON
                            with open(json_path, "w", encoding="utf-8") as f:
                                json.dump(taux_data, f, ensure_ascii=False, indent=4)
                            devise_active = st.session_state.devise_active  # "EUR" ou "USD"

                            # Créer les colonnes si elles n'existent pas
                            for col in ["Taux de change", "Facture HT en devise", "Facture TTC en devise"]:
                                if col not in df_client.columns:
                                    df_client[col] = None

                            # Mapping mois texte → JSON
                            mois_dict = {m: m for m in df_client["Mois"].unique()}

                            # Appliquer le taux pour chaque ligne
                            for idx, row in df_client.iterrows():
                                mois = row["Mois"]
                                taux = taux_data.get(mois_dict[mois], {}).get(devise_active)
                                if taux is not None:
                                    df_client.at[idx, "Taux de change"] = taux
                                    df_client.at[idx, "Facture HT en devise"] = row["Facture HT + NDF"] / taux
                                    df_client.at[idx, "Facture TTC en devise"] = df_client.at[idx, "Facture HT en devise"] * 1.19




                else : 
                        df_client["TAP (DZD)"] = 0.0
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur))
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            import json
                            from pathlib import Path

                            # Chemin du JSON
                            json_path = Path("taux_change.json")

                            # Charger JSON existant ou créer un dictionnaire vide
                            if json_path.exists():
                                with open(json_path, "r", encoding="utf-8") as f:
                                    taux_data = json.load(f)
                            else:
                                taux_data = {}

                            # Exemple : df_rates contient la capture de change
                            # On récupère le mois actuel du système
                            from datetime import datetime
                            mois_courant_num = datetime.now().month
                            mois_list = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                                        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                            mois_courant = mois_list[mois_courant_num - 1]

                            # Récupérer les taux de la capture
                            eur_rate = df_rates.loc[df_rates["Devise"].str.contains("EUR"), "Achat"].values[0]
                            usd_rate = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Mettre à jour le JSON
                            taux_data[mois_courant] = {"EUR": float(eur_rate), "USD": float(usd_rate)}

                            # Écrire le JSON
                            with open(json_path, "w", encoding="utf-8") as f:
                                json.dump(taux_data, f, ensure_ascii=False, indent=4)
                            devise_active = st.session_state.devise_active  # "EUR" ou "USD"

                            # Créer les colonnes si elles n'existent pas
                            for col in ["Taux de change", "Facture HT en devise", "Facture TTC en devise"]:
                                if col not in df_client.columns:
                                    df_client[col] = None

                            # Mapping mois texte → JSON
                            mois_dict = {m: m for m in df_client["Mois"].unique()}

                            # Appliquer le taux pour chaque ligne
                            for idx, row in df_client.iterrows():
                                mois = row["Mois"]
                                taux = taux_data.get(mois_dict[mois], {}).get(devise_active)
                                if taux is not None:
                                    df_client.at[idx, "Taux de change"] = taux
                                    df_client.at[idx, "Facture HT en devise"] = row["Facture HT + NDF"] / taux
                                    df_client.at[idx, "Facture TTC en devise"] = df_client.at[idx, "Facture HT en devise"] * 1.19

                
                # Construire un DataFrame avec toutes les nouvelles colonnes
                new_cols = pd.DataFrame({
                    "Frais remboursement": df_client["Frais remboursement calcule"],
                    "Salaire de base": df_client["Salaire de base calcule"],
                    "Indemnité de panier": df_client["Indemnité de panier calcule"],
                    "Indemnité de transport": df_client["Indemnité de transport calcule"],
                    "Facture TVA": df_client["Facture HT + NDF"] * (1.19 / 100),
                    "Facture TTC": df_client["Facture HT + NDF"] * (1 + 1.19 / 100),
                    "Observation": Observation
                })

                # Fusionner en une seule fois
                df_client = pd.concat([df_client, new_cols], axis=1).copy()

                print("👀 Employés bruts dans df_client :", df_client["Nom"].unique())

                st.write(df_client.head(50)) # On peut encapsuler ton code de calculs dans une fonction
                print(df_client[["Nom","N°"]])
                print(df_client.query("Nom == 'AMALOU'")[["Nom", "N°", "Mois"]])

                # 1. On définit les colonnes fixes (identité employé)
                # Colonnes fixes (identité employé)
                mois_ordre = [
                    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
                ]


                id_cols = ["Nom", "Prénom", "N°",  "Etablissement", "Année"]

                # Colonnes variables (toutes sauf identités + Mois)
                val_cols = [c for c in df_client.columns if c not in id_cols + ["Mois"]]

                # Pivot
                nb_candidats = df_client["N°"].nunique()

                if df_client["N°"].nunique() == 1:
                    # Un seul candidat → regrouper toutes les lignes par mois
                    employe_data = {}
                    for _, row in df_client.iterrows():
                        mois = row["Mois"]
                        for col in df_client.columns:
                            if col not in id_cols + ["Mois"]:
                                employe_data[f"{col}_{mois}"] = row[col]

                    # Réordonner les colonnes par mois_ordre
                    ordered_employe_data = {}
                    # Ajouter d'abord les colonnes statiques
                    for col in id_cols:
                        ordered_employe_data[col] = df_client.iloc[0][col]

                    # Ajouter les colonnes par mois
                    for mois in mois_ordre:
                        for col in val_cols:  # toutes les colonnes variables
                            key = f"{col}_{mois}"
                            if key in employe_data:
                                ordered_employe_data[key] = employe_data[key]

                    employe_data = ordered_employe_data
                else:
                   
                    df_pivot = (
                        df_client
                        .groupby(id_cols + ["Mois"], dropna=False)[val_cols]
                        .sum(min_count=1)
                        .unstack(fill_value=None)
                    )

                    # Aplatir MultiIndex colonnes
                    df_pivot.columns = [f"{val}_{mois}" for val, mois in df_pivot.columns]

                    # ✅ Forcer la présence de tous les employés
                    tous_employes = df_client[id_cols].drop_duplicates().set_index(id_cols)

                    df_pivot = (
                        df_pivot
                        .reindex(pd.MultiIndex.from_frame(tous_employes.reset_index()))
                        .reset_index()
                    )
                    # Réordonner les colonnes par mois_ordre
                    colonnes_identite = id_cols
                    colonnes_mois = []
                    for mois in mois_ordre:
                        colonnes_mois.extend([c for c in df_pivot.columns if c.endswith(f"_{mois}")])

                    df_pivot = df_pivot[colonnes_identite + colonnes_mois]

                    print(df_pivot[["Nom","N°"]])

                # ---- Récupération tables ----
                df_paie = df_client           # votre table paie
                df_travel = pd.DataFrame(supabase.table("g_d").select("*").execute().data)
                df_paie["Mois"] = df_paie["Mois"].str.strip().str.capitalize()
                df_travel["Mois"] = df_travel["Mois"].str.strip().str.capitalize()

                print(df_travel.columns)
                print(df_travel.head())
                print(df_paie["Mois"].unique())
                print(df_travel["Mois"].unique())
                print(df_travel["N°"].head())

                # ---- Fusion ----
                df_merged = df_paie.merge(df_travel, on=["N°", "Mois"], how="left")
                # st.write("Colonnes df_merged :", df_merged.columns.tolist())
                df_merged = df_merged.rename(columns={
                    "Nom_x": "Nom",
                    "Prénom_x": "Prénom",
                    "Etablissement_x": "Etablissement",
                    "Année_x": "Année",
                    "Mois_x": "Mois",

                })
                colonnes_y = [c for c in df_merged.columns if c.endswith("_y")]
                df_merged = df_merged.drop(columns=colonnes_y)
                df_merged.columns = df_merged.columns.str.strip().str.replace("_", " ")

                print(df_merged.head(30))
                print(df_merged.columns)

                # ---- Colonnes manquantes (sécurité) ----
                colonnes_travel = [
                    "Travel expenses M segment",
                    "Travel expenses C segment",
                    "Allowance M segment",
                    "Allowance C segment"
                ]
                
                for col in colonnes_travel:
                    if col not in df_merged.columns:
                        df_merged[col] = None
                for col in colonnes_travel:
                    df_merged[col] = pd.to_numeric(df_merged[col], errors="coerce").fillna(0)
                df_merged["Total Travel expenses"] = df_merged["Travel expenses M segment"] + df_merged["Travel expenses C segment"]
                df_merged["Total Allowance"] = df_merged["Allowance M segment"] + df_merged["Allowance C segment"]
                TAXE= 0.05
                df_merged["Total Travel expenses before VAT"] = df_merged["Total Travel expenses"]+(df_merged["Total Travel expenses"]*TAXE)
                df_merged["Total Travel expenses including VAT"] = df_merged["Total Travel expenses before VAT"]+(df_merged["Total Travel expenses before VAT"]*0.19)
                df_merged["Total Allowance before VAT"]= df_merged["Total Allowance"]+(df_merged["Total Allowance"]*TAXE)
                df_merged["Total Allowance Including VAT"] = df_merged["Total Allowance before VAT"]+(df_merged["Total Allowance before VAT"]*0.19)

                import json

                with open("taux_change.json", "r", encoding="utf-8") as f:
                    taux_data = json.load(f)
                # Colonnes à créer si elles n'existent pas déjà
                for col in ["Taux de change", "Total Travel expenses en devise", "Total Allowance en devise"]:
                    if col not in df_merged.columns:
                        df_merged[col] = None

                # Dictionnaire mois texte
                mois_dict = {m: m for m in df_merged["Mois"].unique()}

                # Application du taux pour chaque ligne
                for idx, row in df_merged.iterrows():
                    mois = row["Mois"]
                    taux = taux_data.get(mois_dict[mois], {}).get("USD")


                    if taux:
                        df_merged.at[idx, "Taux de change"] = taux

                        # Conversion en devise
                        df_merged.at[idx, "Total Travel expenses en devise"] = row["Total Travel expenses before VAT"] / taux
                        df_merged.at[idx, "Total Allowance en devise"] = row["Total Allowance before VAT"] / taux

                df_merged["Taxe"]=TAXE
                colonnes_travel = [
                    "Travel expenses M segment",
                    "Travel expenses C segment",
                    "Allowance M segment",
                    "Allowance C segment",
                    "Total Travel expenses",
                    "Total Allowance",
                    "Taxe",
                    "Total Travel expenses before VAT",
                    "Total Travel expenses including VAT",
                    "Total Allowance before VAT",
                    "Total Allowance Including VAT",
                    "Taux de change",
                    "Total Travel expenses en devise",
                    "Total Allowance en devise"

                ]
                colonnes_allowance = [
                    "Allowance M segment",
                    "Allowance C segment",
                    "Total Allowance",
                    "Taxe",
                    "Total Allowance before VAT",
                    "Total Allowance Including VAT",
                    "Taux de change",
                    "Total Allowance en devise"

                ]
                df_travel_pivot = (
                    df_merged
                    .groupby(["N°", "Mois"], dropna=False)[colonnes_travel]
                    .sum(min_count=1)
                    .unstack(fill_value=None)
                )
                df_allowance_pivot = (
                    df_merged
                    .groupby(["N°", "Mois"], dropna=False)[colonnes_allowance]
                    .sum(min_count=1)
                    .unstack(fill_value=None)
                )
                df_travel_pivot.columns = [f"{val}_{mois}" for val, mois in df_travel_pivot.columns]
                df_allowance_pivot.columns = [f"{val}_{mois}" for val, mois in df_allowance_pivot.columns]
                df_travel_pivot = df_travel_pivot.reset_index()
                df_allowance_pivot = df_allowance_pivot.reset_index()

                # ---- Génération Excel ----
                

                # 📥 Génération et téléchargement Excel
                # --------------------------------------
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_client.to_excel(writer, index=False, sheet_name='Calculs')
                    workbook = writer.book
                    worksheet = writer.sheets['Calculs']

                    # Style entête
                    header_format = workbook.add_format({
                        'bold': True, 'text_wrap': True, 'valign': 'middle',
                        'align': 'center', 'fg_color': '#0a5275',
                        'font_color': 'white', 'border': 1
                    })
                    for col_num, value in enumerate(df_client.columns.values):
                        worksheet.write(0, col_num, value, header_format)

                    # Ajuster largeur colonnes
                    for i, col in enumerate(df_client.columns):
                        col_width = max(df_client[col].astype(str).map(len).max(), len(col)) + 2
                        worksheet.set_column(i, i, col_width)

                st.download_button(
                    label="📊 Télécharger les résultats en Excel",
                    data=output.getvalue(),
                    file_name=f"{st.session_state.selected_client}_calculs.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

               
                # 📥 Génération et téléchargement PDF par employé
                # ------------------------------------------------
                st.markdown("### 📥 Télécharger la facture PDF par employé")

                # 📌 Nouveau fichier par client
                
                # Nom du client
                client_name = str(df_client["Etablissement"].iloc[0]).strip()
                fichier_client = f"{client_name}_factures.xlsx"

                wb = Workbook()
                wb.remove(wb.active)  # supprimer la feuille vide

                if nb_candidats == 1:
                    # Un seul employé → créer UNE seule feuille consolidée
                    nom = df_client.iloc[0]["Nom"].replace(" ", "_")
                    matricule = str(df_client.iloc[0]["N°"])
                    
                    ws = wb.create_sheet(title=f"{matricule}_{nom}")
                    generer_facture_excel_sheet(employe_data, ws, wb)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)

                    # --- Boucle UNE FOIS par employé ---
                    for employe_id in df_merged["N°"].unique():

                        # récupérer toutes les lignes de cet employé
                        df_emp = df_merged[df_merged["N°"] == employe_id]

                        # infos fixes
                        info = df_emp.iloc[0].to_dict()

                        # pivot salaire
                        ligne_pivot = df_pivot[df_pivot["N°"] == employe_id].iloc[0].to_dict()

                        # pivot travel
                        ligne_travel = df_travel_pivot[df_travel_pivot["N°"] == employe_id].iloc[0].to_dict()

                        # FUSION TOTALE
                        employe_data = {**info, **ligne_pivot, **ligne_travel}

                        # nom de la feuille
                        matricule = str(info["N°"])
                        nom = info["Nom"].replace(" ", "_")

                        ws = wb.create_sheet(title=f"{matricule}_{nom}")
                        create_recap_sheet(wb, df_pivot, df_travel_pivot, df_allowance_pivot, taux_data, devise="USD")
                        # génération Excel
                        generer_facture_excel_sheet(employe_data, ws, wb)


                        

                # Sauvegarde
                wb.save(fichier_client)

                # Télécharger
                with open(fichier_client, "rb") as f:
                    st.download_button(
                        label=f"📘 Télécharger factures client {client_name}",
                        data=f.read(),
                        file_name=fichier_client,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                drive_file_id = upload_to_drive(
                    fichier_client,  # <-- utiliser fichier_client ici
                    client_name=row.get("Etablissement", "Inconnu") if nb_candidats > 1 else client_name,
                    root_folder_id="0AM1AktJToIM1Uk9PVA"
                )

                # Supprimer le fichier local
                os.remove(fichier_client)

            else:
                st.warning("⚠️ Aucun employé trouvé pour ce client ")
        else:
            st.info("Veuillez d'abord téléverser le fichier récapitulatif global dans la barre latérale.")
        