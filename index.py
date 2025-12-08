import requests
import streamlit as st
import pandas as pd
import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
import bcrypt
import json
import os
import hashlib
import base64
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
import calendar
import math
import numpy as np
import io
from decimal import Decimal
from pathlib import Path
from rapidfuzz import process, fuzz
from reportlab.platypus import Image
from supabase import create_client, Client
from PIL import Image  
import pytesseract
import easyocr
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.errors import HttpError
import os, pickle
import webbrowser
import threading
import time
import pkg_resources

# 📷 Afficher un logo
st.set_page_config(
    page_title="Gestion de la Facturation",
    page_icon="logo2.png",  # chemin local ou URL
    layout="wide"
)
st.markdown("""
<style>
/* 🖼 Logo centré */
.logo-container {
    display: flex;
    justify-content: center;
    align-items: center;
    margin-bottom: 20px;
}

.logo {
    width: 200px;
    height: auto;
}

/* 🌈 Arrière-plan personnalisé + forcer mode sombre */
html, body, .stApp {
    background: #1d2e4e !important;
    font-family: 'Segoe UI', sans-serif;
    color-scheme: dark !important; /* Empêche l'inversion automatique */
    color: white !important;
}

/* 🖍️ Titre centré et coloré */
.main > div > div > div > div > h1 {
    text-align: center;
    color: #00796B !important;
}

/* 🧼 Nettoyage des bordures Streamlit */
.css-18e3th9 {
    padding: 1rem 0.5rem;
}

/* 🎨 Sidebar */
section[data-testid="stSidebar"] {
    background-color: #1f3763 !important;
    color: white !important;
}

section[data-testid="stSidebar"] .css-1v3fvcr {
    color: white !important;
}

/* 🌈 Titres dans la sidebar */
section[data-testid="stSidebar"] h1, 
section[data-testid="stSidebar"] h2, 
section[data-testid="stSidebar"] h3 {
    color: #e01b36 !important;
}

/* 🎨 Barre supérieure */
header[data-testid="stHeader"] {
    background-color: #06dbae !important;
    color: white !important;
}

/* 🧪 Supprimer la transparence */
header[data-testid="stHeader"]::before {
    content: "";
    background: none !important;
}

/* 📱 Correction mobile : forcer couleurs partout */
h1, h2, h3, p, span, label {
    color: white !important;
}

/* 🔵 Boutons bleu foncé forcés */
.stButton button {
    background-color: #2b2c36 !important; /* Bleu foncé */
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.5rem 1rem !important;
    font-weight: bold !important;
    -webkit-appearance: none !important; /* Évite style par défaut mobile */
    appearance: none !important;
}

.stButton button:hover {
    background-color: #43444e !important; /* Bleu plus clair au survol */
    color: white !important;
}
</style>
""", unsafe_allow_html=True)



# 🖼️ Ajouter un logo (remplacer "logo.png" par ton fichier ou une URL)
with open("logo.png", "rb") as image_file:
    encoded = base64.b64encode(image_file.read()).decode()

st.markdown(
    f"""
    <div class="logo-container">
        <img class="logo" src="data:image/png;base64,{encoded}">
    </div>
    """,
    unsafe_allow_html=True
)
st.markdown(
    "<h1 style='text-align: center;'>Bienvenue sur l'application de calcul des factures</h1>",
    unsafe_allow_html=True
)
def nettoyer_colonne(df, col):
    series = (
        df[col]
        .astype(str)
        .str.replace(r"[^\d,.-]", "", regex=True)
        .str.replace(",", ".", regex=False)
        .replace("", "0")
    )

    # Vérifie quelles valeurs ne sont pas convertibles
    erreurs = []
    for val in series.unique():
        try:
            float(val)
        except ValueError:
            erreurs.append(val)

    if erreurs:
        print(f"⚠️ Colonne '{col}' contient des valeurs non convertibles : {erreurs[:20]}")

    # Conversion sécurisée
    return pd.to_numeric(series, errors="coerce").fillna(0.0)



from math import floor
def calcul_base(row):
    salaire = Decimal(str(row["Salaire de base calcule"]))
    prime = Decimal(str(row["Prime mensuelle (Barème) (DZD)"]))
    panier = Decimal(str(row["Indemnité de panier (DZD)"]))
    transport  = Decimal(str(row["Indemnité de transport (DZD)"]))

    return floor((salaire + prime) - ((salaire + prime) * Decimal("0.09")) + (panier+transport))
def get_valeur(col_base, col_nouveau):
    if col_nouveau in df_client.columns:
        return nettoyer_colonne(df_client, col_nouveau).where(
            lambda x: x != 0, nettoyer_colonne(df_client, col_base)
        )
    else:
        return nettoyer_colonne(df_client, col_base)
import unicodedata
import re
import difflib

def normalize_text(s):
    """Normalise pour comparaison: enlever diacritiques, NBSP, espaces multiples, lower."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()

def trouver_client_robuste(client_name, df, debug=False):
    """
    Filtrage tolérant pour retrouver les lignes d'un client dans df.
    Stratégies (en cascade):
      1) égalité normalisée
      2) contenant (normalized contains)
      3) fuzzy close match (difflib) sur la liste d'établissements uniques
    Retourne df_filtré (copie). Si debug=True, renvoie aussi info dict.
    """
    if df is None or df.empty:
        return (pd.DataFrame(), {"reason": "empty_df"}) if debug else pd.DataFrame()

    df2 = df.copy()
    df2["Etablissement_norm"] = df2["Etablissement"].astype(str).apply(normalize_text)
    client_norm = normalize_text(client_name)

    # 1) égalité normalisée
    mask_eq = df2["Etablissement_norm"] == client_norm
    df_eq = df2[mask_eq].copy()
    if not df_eq.empty:
        if debug:
            return df_eq, {"method": "exact_norm", "count": len(df_eq)}
        return df_eq

    # 2) contains normalisé (client substring of Etab_norm or vice-versa)
    mask_contains = df2["Etablissement_norm"].str.contains(re.escape(client_norm), na=False)
    df_contains = df2[mask_contains].copy()
    if not df_contains.empty:
        if debug:
            return df_contains, {"method": "contains_etab", "count": len(df_contains)}
        return df_contains

    # also try client contained in etab or etab contained in client (reverse)
    mask_rev = df2["Etablissement_norm"].apply(lambda x: client_norm in x or x in client_norm)
    df_rev = df2[mask_rev].copy()
    if not df_rev.empty:
        if debug:
            return df_rev, {"method": "contains_rev", "count": len(df_rev)}
        return df_rev

    # 3) fuzzy matching on unique values
    uniques = sorted(df2["Etablissement_norm"].dropna().unique().tolist())
    close = difflib.get_close_matches(client_norm, uniques, n=3, cutoff=0.7)
    if close:
        # prendre le premier best match
        best = close[0]
        df_close = df2[df2["Etablissement_norm"] == best].copy()
        if debug:
            return df_close, {"method": "fuzzy", "match": best, "close_candidates": close, "count": len(df_close)}
        return df_close

    # Aucun résultat
    if debug:
        # retourner un petit aperçu des candidats potentiels pour aider
        sample_uniques = uniques[:30]
        return pd.DataFrame(), {"method": "no_match", "sample_candidates": sample_uniques}

    return pd.DataFrame()

def fetch_all_data(table_name, batch_size=1000):
    all_data = []
    last_id = 0
    
    while True:
        response = (
            supabase.table(table_name)
            .select("*")
            .order("id")
            .gte("id", last_id + 1)
            .limit(batch_size)
            .execute()
        )
        
        if not response.data:
            break
        
        all_data.extend(response.data)
        
        # mettre à jour le dernier id récupéré
        last_id = max(item["id"] for item in response.data)
        
        if len(response.data) < batch_size:
            break
    
    return pd.DataFrame(all_data)

from reportlab.platypus import Image, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors


def calcul_joursstc_ouvres(row):
    val = row["Date du dernier jour travaillé"]
    nbr_jours_stc = row["Nbr jours STC (jours)"]
    
    date_debut = pd.to_datetime(val, dayfirst=True, errors="coerce")
    if pd.isna(date_debut) or nbr_jours_stc == 0:
        return 0

    start = (date_debut + pd.Timedelta(days=1)).normalize()
    end = start + pd.Timedelta(days=nbr_jours_stc - 1)

    # Générer toutes les dates de la période
    toutes_les_dates = pd.date_range(start=start, end=end, freq="D")
    
    # Filtrer les jours ouvrés (exclure vendredi=4 et samedi=5)
    jours_ouvres = sum(1 for date in toutes_les_dates if date.weekday() not in [4, 5])
    
    return jours_ouvres

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account

def authenticate_drive():
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    service = build("drive", "v3", credentials=creds)

    try:
        # Test: lister quelques fichiers accessibles
        results = service.files().list(pageSize=5, fields="files(id, name)").execute()
        # st.success("✅ Connexion Drive réussie")
        # st.write(results.get("files", []))
    except Exception as e:
        st.error(f"❌ Erreur connexion Drive: {e}")

    return service



def get_or_create_folder(service, folder_name, parent_id=None, drive_id=None):
    """
    Vérifie si un dossier existe dans Google Drive (y compris Drive partagé), sinon le crée.
    Retourne l'ID du dossier.
    """
    # Requête pour chercher le dossier
    query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and trashed=false"
    if parent_id:
        query += f" and '{parent_id}' in parents"

    results = service.files().list(
        q=query,
        fields="files(id, name)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    items = results.get("files", [])

    if items:
        return items[0]["id"]  # ✅ Dossier trouvé

    # Sinon, création du dossier
    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder"
    }
    if parent_id:
        metadata["parents"] = [parent_id]
    if drive_id:  # obligatoire pour Drive partagé si parent_id est la racine
        metadata["driveId"] = drive_id

    folder = service.files().create(
        body=metadata,
        fields="id",
        supportsAllDrives=True
    ).execute()
    return folder["id"]


from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
import os

def upload_to_drive(file_path, client_name, root_folder_id=None, drive_id=None):
    service = authenticate_drive()

    # 1️⃣ Vérifier/créer le dossier client
    folder_id = get_or_create_folder(service, client_name, parent_id=root_folder_id, drive_id=drive_id)

    # Nom du fichier (nom local)
    file_name = os.path.basename(file_path)

    # 2️⃣ Vérifier si un fichier avec le même nom existe déjà
    query = f"name='{file_name}' and '{folder_id}' in parents and trashed=false"
    results = service.files().list(
        q=query,
        fields="files(id, name)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    existing_files = results.get("files", [])

    if existing_files:
        file_id = existing_files[0]["id"]
        print(f"♻️ Mise à jour du fichier existant : {file_name} ({file_id})")

        media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        file = service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True
        ).execute()
    else:
        file_metadata = {"name": file_name, "parents": [folder_id]}
        if drive_id:  # utile si on est dans un Drive partagé
            file_metadata["driveId"] = drive_id

        media = MediaFileUpload(file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        try:
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields="id",
                supportsAllDrives=True
            ).execute()
        except HttpError as e:
            st.error("⚠️ Erreur API Google Drive")
            st.code(e.content.decode("utf-8") if hasattr(e, "content") else str(e))
            raise

    print(f"✅ Fichier disponible dans {client_name} : {file_name} ({file['id']})")
    return file["id"]


def generer_facture_excel_sheet(employe_dict, ws, wb, logos_folder="facturation_app/Logos"):
    # 📌 Styles
    
#     # 📌 Styles
    header_font = Font(bold=True, size=14, color="000000")
    normal_font_black = Font(size=11, color="000000")
    normal_font_white = Font(size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    COL_OFFSET = 4
    
    # 📌 Mapping couleurs
    color_map = {
        "Base cotisable": "9fc5e8", "Retenue CNAS employé": "9fc5e8",
        "Base imposable au baréme": "9fc5e8", "IRG barème": "9fc5e8",
        "Base imposable 10%": "9fc5e8", "IRG 10%": "9fc5e8",
        "Salaire brut": "9fc5e8", "CNAS employeur": "9fc5e8",
        "Cotisation œuvre sociale": "9fc5e8", "Taxe formation": "9fc5e8",
        "Taxe formation et os": "9fc5e8", 
        "Coût congé payé": "9fc5e8", "Taux complément santé (DZD)": "9fc5e8",
        "Fees etalent": "9fc5e8", "TAP": "9fc5e8",
        "Salaire net": "25488e", "Masse salariale": "25488e", "Coût salaire": "25488e",
        "Facture HT": "e11b36","Facture TVA": "284052", "Facture TTC": "284052",
    }
    white_text_lines = {"Salaire net", "Masse salariale", "Coût salaire",
                        "Facture HT", "Facture TVA", "Facture TTC"}
    

    # ---------------------- LOGO ----------------------
    etablissement = str(employe_dict.get("Etablissement", "")).strip()
    logo_path = os.path.join(logos_folder, f"{etablissement}.png")

    if os.path.exists(logo_path):
        try:
            logo = XLImage(logo_path)
            logo.width = 350
            logo.height = 120
            ws.add_image(logo, f"{get_column_letter(COL_OFFSET+4)}1")
        except:
            pass

    # ---------------------- INFOS EMPLOYE ----------------------
    infos_employe = [
        ["Nom:", employe_dict.get("Nom", "")],
        ["Prénom:", employe_dict.get("Prénom", "")],
        ["Année:", employe_dict.get("Année", "")],
        ["Titre du poste:", employe_dict.get("Titre du poste", "")],
        ["Durée CDD:", employe_dict.get("Durée du CDD (Mois)", "")],
        ["Établissement:", etablissement]
    ]

    for i, (label, value) in enumerate(infos_employe, start=3):
        ws.cell(row=i, column=COL_OFFSET, value=label).font = Font(bold=True)
        ws.cell(row=i, column=COL_OFFSET+1, value=value).font = normal_font_black
    # 📌 Catégorisation clients
        clients_simples = ["Abbott", "Samsung"]
        client_sante = ["Siemens", "Healthineers","Siemens Energy", "Siemens Healthineers Oncology",
                        "Tango","CCIS ex SOGEREC","JTI",
                        "Wilhelmsen", "IPSEN", "LG","INTERTEK","Castel el Djazair"]
        client_os = ["Maersk", "Henkel"]
        client_change = ["Epson"]
        client_change_phone = ["Cahors"]
        client_ndf = ["Syngenta"]
        client_gd = ["G+D"]
        client_PMI=["Philip Morris International","Roche"]
    # ---------------------- EXTRACTION DES MOIS ----------------------
    mois_data = {}
    for key, value in employe_dict.items():
        if "_" in key:
            ligne_nom, mois_nom = key.rsplit("_", 1)
            mois_data.setdefault(mois_nom, {})[ligne_nom] = value

    def is_useful_value(v):
        if v is None:
            return False
        try:
            if isinstance(v, float) and math.isnan(v):
                return False
        except:
            pass
        if isinstance(v, str):
            if v.strip().lower() in ("", "nan", "none", "null"):
                return False
            return True
        return True

    mois_disponibles = [m for m, lignes in mois_data.items()
                        if any(is_useful_value(v) for v in lignes.values())]

    # ---------------------- TABLEAU ----------------------
    def generer_tableau(start_row, titre, lignes):
        # Titre tableau
        ws.merge_cells(start_row=start_row, start_column=COL_OFFSET, 
                       end_row=start_row, end_column=COL_OFFSET+len(mois_disponibles))
        titre_cell = ws.cell(row=start_row, column=COL_OFFSET, value=titre)
        titre_cell.font = Font(bold=True, size=13, color="000000")
        titre_cell.alignment = center_alignment
        start_row += 1

        # En-tête
        ws.cell(row=start_row, column=COL_OFFSET, value="Éléments").font = header_font
        ws.cell(row=start_row, column=COL_OFFSET).fill = header_fill
        ws.cell(row=start_row, column=COL_OFFSET).alignment = center_alignment
        ws.cell(row=start_row, column=COL_OFFSET).border = border

        for col, mois in enumerate(mois_disponibles, start=1):
            cell = ws.cell(row=start_row, column=COL_OFFSET+col, value=mois)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Lignes
        for row, ligne in enumerate(lignes, start=1):
            current_row = start_row + row
            fill_color = color_map.get(ligne)
            fill_style = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
            font_color = "FFFFFF" if ligne in white_text_lines else "000000"

            cell_titre = ws.cell(row=current_row, column=COL_OFFSET, value=ligne)
            cell_titre.font = Font(bold=True, color=font_color)
            cell_titre.alignment = left_alignment
            cell_titre.border = border
            if fill_style: cell_titre.fill = fill_style

            for col, mois in enumerate(mois_disponibles, start=1):
                val = mois_data.get(mois, {}).get(ligne, " ")
                if isinstance(val, (int, float)):
                    val = f"{val:,.2f}".replace(",", " ").replace(".", ",")
                cell = ws.cell(row=current_row, column=COL_OFFSET+col, value=val)
                cell.font = Font(size=11, color=font_color)
                cell.alignment = center_alignment
                cell.border = border
                if fill_style: cell.fill = fill_style
                elif row % 2 == 0: cell.fill = data_fill

        return current_row + 2  # ligne suivante

    # 📌 Cas spécial G+D → 3 tableaux
    if etablissement in client_gd:
        start_row = 10
        start_row = generer_tableau(start_row, "Récapitulatif salarial",
            ["Salaire de base","Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Prime mensuelle", "Frais de remboursement (Véhicule) (DZD)",
             "Salaire net","Coût salaire", "Facture HT","Facture TVA", "Facture TTC"])
#         # start_row = generer_tableau(start_row, "Travel Expenses",
#         #     ["Travel expenses M segment", "Travel expenses C segment"])
#         # start_row = generer_tableau(start_row, "Allowance",
#         #     ["Allowance M segment", "Allowance C segment"])

#     # 📌 Tous les autres clients
    else:
        if etablissement in clients_simples:
            lignes = ["Salaire de base", "Prime mensuelle", "Prime exeptionnelle (10%) (DZD)",
                      "Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Frais remboursement",
                      "Base cotisable", "Retenue CNAS employé", "Base imposable au baréme", "IRG barème",
                      "Base imposable 10%", "IRG 10%", "Salaire net", "Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation", "Masse salariale", 
                      "Coût congé payé","Coût salaire","Fees etalent", "Facture HT","Facture TVA", "Facture TTC"]
        elif etablissement in client_sante:
            lignes = ["Salaire de base", "Prime mensuelle", "Prime exeptionnelle (10%) (DZD)",
                      "Prime vestimentaire (DZD)", "Indemnités Non Cotisable - Mensuelle | Panier, Transport", 
                      "Frais remboursement","Base cotisable", "Retenue CNAS employé", "Base imposable au baréme",
                      "IRG barème","Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation", "Masse salariale", 
                      "Coût congé payé","Taux complément santé (DZD)","Coût salaire","Fees etalent", 
                      "Facture HT","Facture TVA", "Facture TTC"]
        elif etablissement in client_os:
            lignes = ["Salaire de base", "Prime mensuelle", "Prime exeptionnelle (10%) (DZD)",
                      "Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Frais remboursement",
                      "Base cotisable", "Retenue CNAS employé", "Base imposable au baréme", "IRG barème",
                      "Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Taxe formation et os","Coût congé payé","Taux complément santé (DZD)","Coût salaire",
                      "Fees etalent","TAP", "Facture HT","Facture TVA", "Facture TTC"]
        elif etablissement in client_change:
            lignes = ["Salaire de base", "Prime mensuelle", "Prime exeptionnelle (10%) (DZD)",
                      "Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Frais remboursement",
                      "Base cotisable", "Retenue CNAS employé", "Base imposable au baréme","IRG barème",
                      "Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation", "Masse salariale", 
                      "Coût congé payé",
                      "Coût salaire","Fees etalent", "Facture HT","Facture TVA","Facture TTC","Facture HT en devise","Facture TTC en devise"]
        elif etablissement in client_change_phone:
            lignes = ["Salaire de base", "Prime mensuelle", "Prime exeptionnelle (10%) (DZD)",
                      "Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Frais remboursement",
                      "Base cotisable", "Retenue CNAS employé", "Base imposable au baréme","IRG barème",
                      "Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation", "Masse salariale", 
                      "Coût congé payé","Coût salaire","Fees etalent", "Facture HT", "Facture TVA","Facture TTC","Facture HT en devise","Facture TTC en devise"]
        elif etablissement in client_ndf:
            lignes = ["Salaire de base", "Prime mensuelle", "Prime exeptionnelle (10%) (DZD)",
                      "Prime vestimentaire (DZD)","Indemnités Non Cotisable - Mensuelle | Panier, Transport", 
                      "Frais remboursement","Base cotisable", "Retenue CNAS employé", "Base imposable au baréme",
                      "IRG barème","Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation", "Masse salariale", 
                      "Coût congé payé","Coût salaire","Fees etalent", "Facture HT","Facture TVA", "Facture TTC"]
        elif etablissement in client_PMI:
            lignes = ["Salaire de base", "Prime mensuelle (Barème) (DZD)", "Prime exeptionnelle (10%) (DZD)",
                      "Prime vestimentaire (DZD)", "Indemnités Non Cotisable - Mensuelle | Panier, Transport", 
                      "Frais remboursement","Base cotisable", "Retenue CNAS employé", "Base imposable au baréme",
                      "IRG barème","Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation", "Masse salariale", 
                      "Coût congé payé","Taux complément santé (DZD)","Coût salaire","Fees etalent", 
                      "Facture HT","Facture TVA", "Facture TTC"]
        else:
            lignes = ["Salaire de base","IFSP (20% du salaire de base)", "Prime mensuelle (Barème) (DZD)",  
                      "Prime exeptionnelle (10%) (DZD)", "Indemnité de panier","indémnité Véhicule",  
                      "Indemnités Non Cotisable - Mensuelle | Panier, Transport", "Frais remboursement",
                      "Base cotisable", "Retenue CNAS employé", "Base imposable au baréme","IRG barème",
                      "Base imposable 10%", "IRG 10%","Salaire net","Salaire brut", "CNAS employeur",
                      "Cotisation œuvre sociale", "Taxe formation","Taxe formation et os", "Masse salariale", 
                      "Coût congé payé","Taux complément santé (DZD)","Coût salaire","Fees etalent", 
                      "Facture HT", "Facture TTC"]


    generer_tableau(10, "Récapitulatif salarial", lignes)

    for col in range(COL_OFFSET, COL_OFFSET + len(mois_disponibles) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 35

    ws.freeze_panes = "E1"

    
def calcul_cout_conge(row):
    # Cas 1 : augmentation ou nouveau salaire
    # if row["Augmentation state"] == "Yes" or row["Nouveau Salaire de base (DZD)"] != 0:
        
    #     Masse_salariale = (
    #                 row["Salaire brut"] +
    #                 row["CNAS employeur"] +
    #                 row["Cotisation œuvre sociale"] +
    #                 row["Taxe formation"]
    #             )
    #     cout_conge = (Masse_salariale *(2.5/30))
    # else:
        # Cas 2 : pas d’augmentation → on garde le coût congé payé existant
    cout_conge = row["Coût congé payé"]

    # Ajout régul si "Congé payé"
    # cout_conge += row["Régul"] if row["Base de régul"] == "Congé payé" else 0
    return cout_conge
USERS_FILE = "users.json"
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
# 🔒 Hachage du mot de passe
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# 📝 Inscription
def signup(email, password):
    response = supabase.auth.sign_up({"email": email, "password": password})
    return response

# 🔐 Connexion
def login(email, password):
    response = supabase.auth.sign_in_with_password({"email": email, "password": password})
    return response

# 🚪 Déconnexion
def logout():
    supabase.auth.sign_out()
    st.session_state["authenticated"] = False
    st.session_state["user"] = None

# --- Interface ---
# st.title("🔐 Auth avec Supabase")

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    mode = st.radio("Choisissez :", ["Connexion", "Créer un compte"])

    email = st.text_input("Email")
    password = st.text_input("Mot de passe", type="password")

    if mode == "Créer un compte":
        if st.button("Créer un compte"):
            res = signup(email, password)
            if res.user:
                st.success("✅ Compte créé ! Vérifie ton email.")
            else:
                st.error(f"Erreur : {res}")
    else:
        if st.button("Connexion"):
            res = login(email, password)
            if res.user:
                st.session_state["authenticated"] = True
                st.session_state["user"] = res.user
                st.rerun()
            else:
                st.error("❌ Email ou mot de passe incorrect.")

else:
    with st.sidebar:
        st.success(f"Bienvenue {st.session_state['user'].email} 👋")
        if st.button("Se déconnecter"):
            logout()
            st.rerun()

# 🎉 Application principale ici
# st.title("Bienvenue sur l'application de calcule des factures")



    # 📌 Initialisation
    if "clients" not in st.session_state:
        st.session_state.clients = []
    if "selected_client" not in st.session_state:
        st.session_state.selected_client = None
    if "full_df" not in st.session_state:
        st.session_state.full_df = None
    if "data" not in st.session_state:
        st.session_state.data = {}


    CLIENTS_FILE = Path("clients.json")
    # st.title("👥 Gestion des clients et des employés")

    # 📂 Charger la liste des clients
    if CLIENTS_FILE.exists():
        with open(CLIENTS_FILE, "r", encoding="utf-8") as f:
            clients_list = json.load(f)
    else:
        clients_list = [
        "Abbott", "Samsung", "Henkel", "G+D", "Maersk",
            "Cahors", "Philip Morris International", "Siemens", "Syngenta", "LG",
            "Epson", "EsteL", "JTI", "Siemens Energy", "Wilhelmsen",
            "Healthineers", "Contrat auto-entrepreneur", "Coca Cola", "IPSEN", "SOGEREC","CCIS ex SOGEREC",
            "Roche", "Tango", "VARION","Castel el Djazair","INTERTEK"
        ]
        with open(CLIENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(clients_list, f, ensure_ascii=False, indent=2)
    client_name = st.sidebar.selectbox(
        "Sélectionner un client",
        options=clients_list,
        index=None,  # <-- pas de sélection initiale
        placeholder="— Sélectionner un client —",
        key="client_select",
    )
    st.session_state.clients = clients_list
    st.session_state.selected_client = client_name
    # 📁 Upload du fichier global
    # GLOBAL_CSV = "data_global.csv"
    st.sidebar.subheader("📅 Charger le fichier Récap")
    is_complement = st.sidebar.checkbox("📌 Complémentaire ?", value=False)
    uploaded_csv = st.sidebar.file_uploader("Fichier CSV Récap", type=["csv"], key="csv_recap")
    
    MOIS_MAP = {
    "-janv.-": "Janvier",
    "-févr.-": "Février",
    "-mars-": "Mars",
    "-avr.-": "Avril",
    "-mai-": "Mai",
    "-juin-": "Juin",
    "-juil.-": "Juillet",
    "-août-": "Août",
    "-sept.-": "Septembre",
    "-oct.-": "Octobre",
    "-nov.-": "Novembre",
    "-déc.-": "Décembre",
    }

    if uploaded_csv is not None:
        try:
            raw = uploaded_csv.getvalue()
            raw = raw.replace(b"\xe2\x80\xaf", b" ")

            # Charger CSV
            df_full = pd.read_csv(
                io.BytesIO(raw),
                skiprows=2,
                sep=",",
                decimal=",",
                thousands=" "
            )
            st.session_state.full_df = df_full
            st.write("📄 Contenu CSV importé :")
            st.dataframe(df_full.head())

            # Nettoyer NaN et inf
            df_full = df_full.replace([np.nan, np.inf, -np.inf], 0)

            # Vérifier et normaliser la colonne Mois
            if "Mois" in df_full.columns:
                df_full["Mois"] = df_full["Mois"].map(MOIS_MAP).fillna(df_full["Mois"])
                mois = df_full["Mois"].iloc[0]
            else:
                st.error("❌ La colonne 'Mois' est manquante dans le CSV")
                mois = "Inconnu"

            # Déterminer la table cible
            table_target = "Recap_complément" if is_complement else "Recap"

            if is_complement:
                # Lire la recap existante pour le même mois
                res = supabase.table("Recap").select("*").eq("Mois", mois).execute()
                cols = supabase.table("Recap").select("*").limit(1).execute()
                # st.write("Colonnes réellement présentes :", list(cols.data[0].keys()))

                df_recap = pd.DataFrame(res.data)
                if df_recap.empty:
                    st.warning("⚠ Pas de récap existante pour ce mois, tout sera considéré comme complément.")
                    df_diff = df_full.copy()
                else:
                    # Colonnes numériques à comparer
                    # Identifier les colonnes à comparer
                    cols_to_check = ["Absence (Jour)","Heure Absence(H)",	"Absence Maladie (Jour)",	"Absence Maternité (Jour)",	"Absence Mise à pied (Jour)","Total absence (sur 22 jours)","Heures supp 100% (H)"	,"Heures supp 75% (H)","Heures supp 50% (H)","Jours supp (Jour)"	,"Prime mensuelle (Barème) (DZD)",	"Prime exeptionnelle (10%) (DZD)","Indemnité non cotisable et imposable 10% (DZD)"	,"Rappel indemnité (brut)","Avance NET (DZD)"	,"Indemnité de départ (Net)"	,"Allocation Aid El Adha NET"	,"Allocations exceptionnelles NET"	,"Jours de congé (Jour)"]
                    for col in cols_to_check:
                        df_full[col] = pd.to_numeric(df_full[col], errors="coerce").fillna(0)
                        if not df_recap.empty:
                            df_recap[col] = pd.to_numeric(df_recap[col], errors="coerce").fillna(0)
                    
                    missing_cols = [col for col in cols_to_check if col not in df_recap.columns]
                    if missing_cols:
                        st.warning(f"⚠ Ces colonnes sont absentes dans Recap et seront considérées à 0 : {missing_cols}")
                        for col in missing_cols:
                            df_recap[col] = 0

                    # Merge pour aligner les lignes par Matricule
                    df_merge = df_full.merge(
                        df_recap[["N°"] + cols_to_check],  # ne prendre que les colonnes à comparer
                        on=["N°"], 
                        suffixes=("_new", "_old")
                    )

                    # Calculer les différences uniquement sur les colonnes numériques
                    df_diff = df_merge.copy()
                    for col in cols_to_check:
                        col_new = f"{col}_new"
                        col_old = f"{col}_old"
                        df_diff[col] = df_merge[col_new] - df_merge[col_old]


                    df_diff = df_diff[["N°"] + cols_to_check]
                    df_diff = df_diff[(df_diff[cols_to_check].sum(axis=1) != 0)]
                    # Colonnes non numériques (ex : booléennes)
                    bool_cols = df_full.select_dtypes(include=["bool"]).columns.tolist()
                    for col in bool_cols:
                        col_new = f"{col}_new"
                        col_old = f"{col}_old"
                        if col_new in df_merge.columns and col_old in df_merge.columns:
                            # Convertir en int pour pouvoir soustraire ou utiliser XOR
                            df_diff[col] = df_diff[col_new].astype(int) - df_diff[col_old].astype(int)


                if not df_diff.empty:
                    st.write("⚡ Différences détectées :")
                    st.dataframe(df_diff)

                    # ✅ Bouton pour valider les différences
                    if st.button("Valider et ajouter au mois suivant"):
                        # Déterminer le mois suivant
                        mois_liste = list(MOIS_MAP.values())
                        mois_actuel = MOIS_MAP.get(mois, mois)
                        idx = mois_liste.index(mois_actuel)
                        mois_suivant = mois_liste[(idx + 1) % 12]

                        st.write(f"Mois suivant : {mois_suivant}")
                        df_diff["Mois"] = mois_suivant

                        # Pour chaque ligne dans df_diff
                        for _, row in df_diff.iterrows():
                            matricule = row["N°"]

                            # Vérifier si la ligne existe déjà pour ce matricule et mois_suivant
                            res = supabase.table("Recap").select("*").eq('"N°"', matricule).eq("Mois", mois_suivant).execute()
                            existing = res.data
                            

                            if existing:
                                
                                existing_row = existing[0]

                                update_data = {}

                                # Colonnes numériques à additionner
                                for col in cols_to_check:
                                    val_existing = existing_row.get(col, 0)
                                    val_new = row.get(col, 0)
                                    
                                    try:
                                        val_existing = float(val_existing)
                                    except:
                                        val_existing = 0
                                        
                                    try:
                                        val_new = float(val_new)
                                    except:
                                        val_new = 0
                                        
                                    update_data[col] = val_existing + val_new

                                # Colonnes restantes (non numériques)
                                for col in df_full.columns:
                                    if col not in cols_to_check:
                                        update_data[col] = row.get(col, existing_row.get(col, None))

                                supabase.table("Recap").update(update_data)\
                                    .eq('"N°"', matricule)\
                                    .eq("Mois", mois_suivant)\
                                    .execute()
                            else:
                                for _, row in df_full.iterrows():
                                    matricule = row["N°"]

                                    # Vérifier si ligne existe déjà pour le mois
                                    res = supabase.table("Recap").select("*").eq('"N°"', matricule).eq("Mois", mois).execute()
                                    existing = res.data
                                    update_data = {}

                                    if existing:
                                        existing_row = existing[0]
                                        for col in df_full.columns:
                                            val_new = row.get(col, 0)
                                            val_existing = existing_row.get(col, 0)

                                            # Colonnes numériques → addition
                                            if col in cols_to_check:
                                                try: val_new = float(val_new)
                                                except: val_new = 0
                                                try: val_existing = float(val_existing)
                                                except: val_existing = 0
                                                update_data[col] = val_existing + val_new
                                            else:
                                                update_data[col] = val_new if val_new not in [None, ""] else val_existing

                                        supabase.table("Recap").update(update_data).eq('"N°"', matricule).eq("Mois", mois).execute()
                                    else:
                                        # Nouvelle ligne
                                        res_id = supabase.table("Recap").select("id").order("id", desc=True).limit(1).execute()
                                        start_id = res_id.data[0]["id"] + 1 if res_id.data else 1
                                        row_dict = row.to_dict()
                                        row_dict["id"] = start_id
                                        for col in df_full.columns:
                                            if col not in row_dict:
                                                row_dict[col] = 0 if col in cols_to_check else None
                                        supabase.table("Recap").insert(row_dict).execute()
                        st.success(f"✅ Compléments ajoutés ou mis à jour dans le mois suivant ({mois_suivant})")

                else:
                    st.info("ℹ Aucun complément détecté, aucune différence trouvée.")

            else:
                st.write("🔄 Fusion avec les données existantes...")

                # Colonnes numériques à additionner
                cols_to_check = ["Absence (Jour)", "Heure Absence(H)", "Absence Maladie (Jour)",
                                "Absence Maternité (Jour)", "Absence Mise à pied (Jour)",
                                "Total absence (sur 22 jours)", "Heures supp 100% (H)",
                                "Heures supp 75% (H)", "Heures supp 50% (H)",
                                "Jours supp (Jour)", "Prime mensuelle (Barème) (DZD)",
                                "Prime exeptionnelle (10%) (DZD)",
                                "Indemnité non cotisable et imposable 10% (DZD)",
                                "Rappel indemnité (brut)", "Avance NET (DZD)",
                                "Indemnité de départ (Net)", "Allocation Aid El Adha NET",
                                "Allocations exceptionnelles NET", "Jours de congé (Jour)"]

                # Récupérer le mois déjà en base (inclut compléments)
                existing = supabase.table("Recap").select("*").eq("Mois", mois).execute()
                df_existing = pd.DataFrame(existing.data)

                if df_existing.empty:
                    st.warning("⚠ Aucun récap existant : insertion directe.")
                    res = supabase.table("Recap").select("id").order("id", desc=True).limit(1).execute()
                    start_id = res.data[0]["id"] + 1 if res.data else 1
                    df_full.insert(0, "id", range(start_id, start_id + len(df_full)))
                    supabase.table("Recap").insert(df_full.to_dict(orient="records")).execute()
                    st.success(f"✅ Données enregistrées pour {mois}")
                else:
                    # MERGE
                    df_merge = df_full.merge(df_existing, on="N°", how="left", suffixes=("_new", "_old"))
                    df_final = df_full.copy()

                    # ADDITION DES COLONNES NUMÉRIQUES
                    for col in cols_to_check:
                        new_col = f"{col}_new"
                        old_col = f"{col}_old"

                        if new_col not in df_merge.columns:
                            # pas de suffixe = colonne inexistante → normaliser à 0
                            df_merge[new_col] = 0
                        if old_col not in df_merge.columns:
                            df_merge[old_col] = 0

                        df_final[col] = (
                            pd.to_numeric(df_merge[new_col], errors="coerce").fillna(0)
                            +
                            pd.to_numeric(df_merge[old_col], errors="coerce").fillna(0)
                        )

                    # COLONNES NON NUMÉRIQUES
                    for col in df_full.columns:
                        if col not in cols_to_check and col != "N°":
                            new_col = f"{col}_new"
                            old_col = f"{col}_old"

                            if new_col in df_merge.columns and old_col in df_merge.columns:
                                df_final[col] = df_merge[new_col].combine_first(df_merge[old_col])
                            else:
                                df_final[col] = df_full[col]

                    # UPDATE / INSERT
                    for _, row in df_final.iterrows():
                        supabase.table("Recap")\
                            .update(row.to_dict())\
                            .eq("Mois", mois)\
                            .eq('"N°"', row["N°"])\
                            .execute()

                    st.success(f"✅ Fusion terminée : récap normal + complément conservés pour {mois}")



        except Exception as e:
            st.sidebar.error(f"❌ Erreur : {e}")
    # ➕ Ajouter un nouveau client
    
    st.sidebar.subheader("➕ Ajouter un nouveau client")
    new_client = st.sidebar.text_input("Nom du nouveau client")
    if st.sidebar.button("Ajouter"):
        if new_client and new_client not in st.session_state.clients:
            st.session_state.clients.append(new_client)
            with open(CLIENTS_FILE, "w", encoding="utf-8") as f:
                json.dump(st.session_state.clients, f, ensure_ascii=False, indent=2)
            st.sidebar.success(f"Client '{new_client}' ajouté !")

    # 🗑️ Supprimer un client avec confirmation
    st.sidebar.subheader("🗑️ Supprimer un client")
    client_to_delete = st.sidebar.selectbox("Choisir le client à supprimer", [""] + st.session_state.clients)

    # Variable temporaire pour confirmation
    if "confirm_delete" not in st.session_state:
        st.session_state.confirm_delete = None

    if st.sidebar.button("Supprimer"):
        if client_to_delete and client_to_delete in st.session_state.clients:
            st.session_state.confirm_delete = client_to_delete  # on garde le nom en mémoire

    # Si un client est en attente de confirmation
    if st.session_state.confirm_delete:
        st.warning(f"⚠️ Êtes-vous sûr de vouloir supprimer le client '{st.session_state.confirm_delete}' ?")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("✅ Oui, supprimer"):
                st.session_state.clients.remove(st.session_state.confirm_delete)
                with open(CLIENTS_FILE, "w", encoding="utf-8") as f:
                    json.dump(st.session_state.clients, f, ensure_ascii=False, indent=2)
                st.success(f"Client '{st.session_state.confirm_delete}' supprimé avec succès !")
                st.session_state.confirm_delete = None  # reset
        with col2:
            if st.button("❌ Annuler"):
                st.info("Suppression annulée.")
                st.session_state.confirm_delete = None


    if st.session_state.selected_client:
        st.markdown(f"## 👤 Données des employés pour **{st.session_state.selected_client.strip()}**")

        if st.session_state.full_df is not None:
            # df = st.session_state.full_df.copy()
            

            df = fetch_all_data("Recap")
            # st.write("Nb lignes récupérées :", len(df))
            # st.write("Mois distincts :", sorted(df["Mois"].unique()))

            # ✅ Filtrer par client
            df_client, info = trouver_client_robuste(st.session_state.selected_client, df, debug=True)
            # st.write("🔍 Debug trouver_client_robuste:", info)
            # st.write("✅ Nombre de lignes trouvées pour le client:", len(df_client))

        # ⚠️ Ne pas écraser totalement, stocker tel quel


            if not df_client.empty:
                # ------------------------------------------------
                # Partie calculs et préparation des données
                # ------------------------------------------------
                mois_possibles = [mois.lower() for mois in calendar.month_name if mois]
                colonnes_mois = [col for col in df_client.columns if any(mois in col.lower() for mois in mois_possibles)]
                nb_employes = df_client["N°"].nunique()
                st.success(f"{nb_employes} employés trouvés.")
                source = st.sidebar.radio("📌 Source des taux :", ["Capture d'écran", "OANDA"])
                if "df_rates" not in st.session_state:
                    st.session_state.df_rates = pd.DataFrame(columns=["Devise", "Achat", "Vente"])
                    devise_active = st.sidebar.radio("Choisir la devise :", ["EUR", "USD"], horizontal=True)
                    st.session_state.devise_active = devise_active
                if source == "Capture d'écran":
                    from PIL import Image  
                    MAX_WIDTH = 800
                    uploaded_image = st.sidebar.file_uploader("📷 Charger la capture", type=["png", "jpg", "jpeg"])

                    if uploaded_image is not None:
                        image = Image.open(uploaded_image)

                        # Redimensionner si trop large
                        width, height = image.size
                        if width > MAX_WIDTH:
                            ratio = MAX_WIDTH / width
                            new_size = (MAX_WIDTH, int(height * ratio))
                            image = image.resize(new_size, Image.Resampling.LANCZOS)

                        st.sidebar.image(image, caption="Capture chargée", use_container_width=True)

                        # OCR avec EasyOCR
                        reader = easyocr.Reader(['en', 'fr'])
                        result = reader.readtext(np.array(image), detail=0)  # texte brut
                        lines = [line.strip() for line in result if line.strip()]

                        # st.sidebar.text_area("📄 Texte brut OCR", value="\n".join(lines), height=200)

                        # 🔎 Extraire les blocs Devise + Achat + Vente
                        data = []
                        i = 0
                        while i < len(lines):
                            line = lines[i]

                            # Devise = 3 lettres majuscules
                            if re.match(r"^[A-Z]{3}$", line):
                                devise = line
                                try:
                                    achat = float(lines[i+1].replace(",", "."))
                                    vente = float(lines[i+2].replace(",", "."))
                                    data.append([devise, achat, vente])
                                    i += 3  # avancer de 3 lignes
                                    continue
                                except Exception:
                                    pass
                            i += 1

                        if data:
                            df_rates = pd.DataFrame(data, columns=["Devise", "Achat", "Vente"])
                            st.write("📊 Taux extraits :")
                            st.dataframe(df_rates)

                            # Sauvegarde en session_state
                            st.session_state.df_rates = df_rates

                            # Interface choix devise
                            devise_active = st.sidebar.radio("Choisir la devise :", ["EUR", "USD"], horizontal=True)
                            st.session_state.devise_active = devise_active


                        else:
                            st.warning("⚠️ Aucun taux détecté dans l'image")
                elif source == "OANDA":
                    if "df_rates" not in st.session_state:
                        st.session_state.df_rates = pd.DataFrame(
                            [["EUR", 1.0, 1.0]],  # Valeur par défaut
                            columns=["Devise", "Achat", "Vente"]
                        )
                    st.sidebar.markdown("## 💱 Saisie du taux de change (OANDA manuel)")
                    with st.sidebar.form("oanda_manual_form"):
                        base = st.text_input("Devise de base", "USD").upper().strip()
                        target = st.text_input("Devise cible", "DZD").upper().strip()
                        taux = st.number_input(
                            f"Taux (1 {base} = ? {target})",
                            min_value=0.0,
                            format="%.6f"
                        )
                        submit = st.form_submit_button("Enregistrer")
                    devise_active = st.sidebar.radio("Choisir la devise :", ["USD", "EUR"], horizontal=True)
                    st.session_state.devise_active = devise_active
                    if submit:
                        if taux <= 0:
                            st.error("Le taux doit être supérieur à 0.")
                        else:
                            pair = f"{base}"
                            new_row = {"Devise": pair, "Achat": taux, "Vente": taux}

                            df = st.session_state.df_rates.copy()
                            mask = df["Devise"] == pair
                            if mask.any():
                                df.loc[mask, ["Achat", "Vente"]] = [taux, taux]
                            else:
                                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                            st.session_state.df_rates = df
                            st.sidebar.success(f"Taux {pair} enregistré : {taux}")

                # --- Affichage du tableau ---
                # if not st.session_state.df_rates.empty:
                #     st.markdown("### 📊 Taux enregistrés")
                #     st.dataframe(st.session_state.df_rates.reset_index(drop=True))
                col1, = st.columns(1) 
                with col1:
                    jours_mois = st.number_input("Jours mois", min_value=28.0, max_value=31.0, step=1.0, value=30.0)

                Observation = st.text_input("Observation", value="")

                # 2. Nettoyage des colonnes
                cols_to_float = [
                    "Salaire de base (DZD)", "Prime mensuelle (Barème) (DZD)", "IFSP (20% du salaire de base)",
                    "Prime exeptionnelle (10%) (DZD)", "Frais de remboursement (Véhicule) (DZD)", 
                    "Indemnité de panier (DZD)", "Indemnité de transport (DZD)", "Nouveau Salaire de base (DZD)",
                    "Prime vestimentaire (DZD)", "Nouvelle Indemnité de panier (DZD)",  "Nouvelle Indemnité de transport (DZD)",
                    "Nouvelle Prime mensuelle (DZD)", "Nouveaux Frais de remboursement (Véhicule) (DZD)","Prime vestimentaire (DZD)", "Indémnité Véhicule (DZD)",
                    "Absence (Jour)","Absence Maladie (Jour)","Absence Maternité (Jour)", "Absence Mise à pied (Jour)", "Jours de congé (Jour)",
                    "Heures supp 100% (H)", "Heures supp 75% (H)", "Heures supp 50% (H)", "Jours supp (Jour)","Taux complément santé (DZD)",
                   "Avance NET (DZD)", "Coût congé payé", "Nbr jours STC (jours)",
                    "Jours de congé ouvré(22 jours)","Indemnité non cotisable et imposable 10% (DZD)", "Total absence (sur 22 jours)",
                    "Nouvelle indémnité Véhicule (DZD)","Nouveau IFSP (20% du salaire de base)","Indemnité de départ (Net)",
                    "Allocation Aid El Adha NET"
                ]
              

                


                for col in cols_to_float:
                    if col in df_client.columns:
                        df_client[col] = nettoyer_colonne(df_client, col)
                    else:
                        df_client[col] = 0.0
                col_pourcentage = ["Fees etalent"]

                for col in col_pourcentage:
                    df_client[col] = (
                        df_client[col]
                        .fillna("0")                          
                        .astype(str)
                        .str.replace("%", "", regex=False)
                        .str.replace(",", ".", regex=False)
                        .str.strip()
                        .replace("", "0")
                        .astype(float)
                    )
                
                # df_client["NDF"]= df_client["NDF"].fillna(0)                                 
                # df_client["jours conge ouvres"] = df_client.apply(calcul_jours_ouvres, axis=1)
                
                df_client["jours stc ouvres"] = df_client.apply(calcul_joursstc_ouvres, axis=1)
                absences_total = (
                df_client["Absence (Jour)"]
                + df_client["Absence Maladie (Jour)"]
                + df_client["Absence Maternité (Jour)"]
                + df_client["Absence Mise à pied (Jour)"]
                + df_client["Jours de congé (Jour)"]  # version brute
            )
                print(df_client["Jours de congé ouvré(22 jours)"])
                absences_total22 = (
                    df_client["Total absence (sur 22 jours)"]
                    + df_client["Jours de congé ouvré(22 jours)"]  # version corrigée week-end
                )
                absences_totallg = (
                df_client["Absence (Jour)"]
                + df_client["Absence Maladie (Jour)"]
                + df_client["Absence Maternité (Jour)"]
                + df_client["Absence Mise à pied (Jour)"]
                
            ) 
                
                HEURES_MOIS = 173.33
                
                # 3. Calculs (une seule fois)
                df_client["Salaire de base calcule"] = (get_valeur("Salaire de base (DZD)", "Nouveau Salaire de base (DZD)"))
                df_client["Indemnité de panier calcule"] = get_valeur("Indemnité de panier (DZD)", "Nouvelle Indemnité de panier (DZD)")
                df_client["indémnité Véhicule calcule"] = get_valeur("Indémnité Véhicule (DZD)", "Nouvelle indémnité Véhicule (DZD)")
                df_client["Indemnité de transport calcule"] = get_valeur("Indemnité de transport (DZD)", "Nouvelle Indemnité de transport (DZD)")
                df_client["Prime mensuelle calcule"] = get_valeur("Prime mensuelle (DZD)", "Nouvelle Prime mensuelle (DZD)")
                df_client["IFSP (20% du salaire de base) calcule"] = get_valeur("IFSP (20% du salaire de base)", "Nouveau IFSP (20% du salaire de base)")
                df_client["Frais remboursement calcule"] = get_valeur("Frais de remboursement (Véhicule) (DZD)", "Nouveaux Frais de remboursement (Véhicule) (DZD)")
                # print(df_client["Salaire de base calcule"])
                # df_client["Salaire de base calcule"] = ((df_client["Salaire de base calcule"]/30)*(30-df_client["Nbr jours augmentation"]))+(((df_client["Salaire de base calcule"] * (1 + (df_client["Augmentation"] / 100)))/30 )*df_client["Nbr jours augmentation"])
                # df_client["Salaire de base calcule"] = (df_client["Salaire de base calcule"] * (1 + (df_client["Augmentation"] / 100)))
                df_client["Salaire de base calcule"] += df_client["IFSP (20% du salaire de base) calcule"]
                salaire_journalier = df_client["Salaire de base calcule"] / jours_mois
                df_client["Salaire de base calcule"] = (
                    (df_client["Salaire de base calcule"]
                    - df_client["Salaire de base calcule"] / 30 * (absences_total)
                    + df_client["Salaire de base calcule"] / HEURES_MOIS * (
                        df_client["Heures supp 100% (H)"] * 2
                        + df_client["Heures supp 75% (H)"] * 1.75
                        + df_client["Heures supp 50% (H)"] * 1.5
                    )
                    + (df_client["Jours supp (Jour)"] * salaire_journalier)) 
                    
                )
                
                df_client["IFSP (20% du salaire de base)"] = df_client["IFSP (20% du salaire de base) calcule"]
                # df_client["Salaire de base calcule"] = df_client["Salaire de base calcule"] - (df_client["Salaire de base calcule"]/jours_mois) * df_client["Nbr jours STC (jours)"]
                # Ajout régul seulement si Base de régul == "Salaire de base"
                # df_client["Salaire de base calcule"] += np.where(
                    # df_client["Base de régul"] == "Salaire de base", df_client["Régul"], 0)

                print(absences_total22)
                if df_client["Etablissement"].iloc[0] == "Coca Cola": 
                    df_client["Indemnité de panier calcule"] = (
                    df_client["Indemnité de panier calcule"]
                    - (df_client["Indemnité de panier calcule"] / 26 * absences_total22)
                    + (df_client["Indemnité de panier calcule"] / 26 * (
                        (df_client["Heures supp 100% (H)"] ) / 8
                        + (df_client["Heures supp 75% (H)"] ) / 8
                        + (df_client["Heures supp 50% (H)"] ) / 8
                    ))
                 )
                    df_client["Indemnité de transport calcule"] = (
                    df_client["Indemnité de transport calcule"]
                    - (df_client["Indemnité de transport calcule"] / 26 * absences_total22)
                    + (df_client["Indemnité de transport calcule"] / 26 * (
                        (df_client["Heures supp 100% (H)"] ) / 8
                        + (df_client["Heures supp 75% (H)"] ) / 8
                        + (df_client["Heures supp 50% (H)"] ) / 8
                    ))
                 )
                    df_client["indémnité Véhicule calcule"] = (
                    df_client["indémnité Véhicule calcule"]
                    - (df_client["indémnité Véhicule calcule"] / 26 * absences_total22)
                    + (df_client["indémnité Véhicule calcule"] / 26 * (
                        (df_client["Heures supp 100% (H)"] ) / 8
                        + (df_client["Heures supp 75% (H)"] ) / 8
                        + (df_client["Heures supp 50% (H)"] ) / 8
                    ))
                 )
              
                    df_client["Indemnitésomme"]= df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] + df_client["Prime vestimentaire (DZD)"] + df_client["indémnité Véhicule calcule"]+df_client["Frais remboursement calcule"]
                    df_client["Indemnité 22jours"] = df_client["Indemnitésomme"]
                    # print(df_client["Indemnité 22jours"])
                else:
                    df_client["Indemnitésomme"]= df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] + df_client["Prime vestimentaire (DZD)"] + df_client["indémnité Véhicule calcule"]
                    # st.write(df_client["Indemnitésomme"].head(133))
                    df_client["Indemnité 22jours"] = (
                        df_client["Indemnitésomme"]
                        - (df_client["Indemnitésomme"] / 22 * (absences_total22 ))
                        + (df_client["Indemnitésomme"] / 22 * (
                            (df_client["Heures supp 100% (H)"] * 2) / 8
                            + (df_client["Heures supp 75% (H)"] * 1.75) / 8
                            + (df_client["Heures supp 50% (H)"] * 1.5) / 8
                        ))
                    )
                   
                    # df_client["Indemnité 22jours"]= df_client["Indemnité 22jours"] - ((df_client["Indemnité 22jours"]/22) * df_client["Nbr jours STC (jours)"])
                # st.write(df_client["Indemnité 22jours"].head(133))
                if df_client["Etablissement"].iloc[0] == "LG":
                    df_client["Indemnitésomme"]= df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] + df_client["Prime vestimentaire (DZD)"] + df_client["indémnité Véhicule calcule"]+df_client["Avance NET (DZD)"] 
                    
                    df_client["Indemnité 22jours"] = (
                        df_client["Indemnitésomme"]
                        - (df_client["Indemnitésomme"] / 30 * df_client["Total absence (sur 22 jours)"])
                        
                    )
                    df_client["Salaire de base calcule"] = (
                        df_client["Salaire de base (DZD)"]
                        + (df_client["Salaire de base (DZD)"] / HEURES_MOIS * (
                        + df_client["Heures supp 100% (H)"] * 2
                        + df_client["Heures supp 75% (H)"] * 1.75
                        + df_client["Heures supp 50% (H)"] * 1.5))
                        - df_client["Salaire de base (DZD)"] / 30 * absences_totallg 
                    )
                    df_client["Salaire de base calcule"] = df_client["Salaire de base calcule"] + df_client["Prime mensuelle (Barème) (DZD)"] 
                    df_client["Base cotisable"] = df_client["Salaire de base calcule"] + df_client["Prime exeptionnelle (10%) (DZD)"]  + df_client["Indemnité non cotisable et imposable 10% (DZD)"] 
                    df_client["Base cotisable"] = (
                            
                         df_client["Salaire de base calcule"] + df_client["Prime mensuelle (Barème) (DZD)"]  + df_client["Indemnité non cotisable et imposable 10% (DZD)"]+ df_client["Prime exeptionnelle (10%) (DZD)"]
                        )
                else:
                    df_client["Base cotisable"] = (
                            
                         df_client["Salaire de base calcule"] + df_client["Prime mensuelle (Barème) (DZD)"]  + df_client["Indemnité non cotisable et imposable 10% (DZD)"]
                        )
                df_client["indémnité Véhicule"] = df_client["indémnité Véhicule calcule"]
                if df_client["Etablissement"].iloc[0] == "LG" :
                    df_client["Base imposable 10%"] = df_client["Indemnité non cotisable et imposable 10% (DZD)"] * 0.91 + df_client["Allocation Aid El Adha NET"]
                else:
                    df_client["Base imposable 10%"] = df_client["Prime exeptionnelle (10%) (DZD)"] * 0.91

                df_client["Retenue CNAS employé"] = df_client["Base cotisable"] * 0.09
                if df_client["Etablissement"].iloc[0] == "Henkel": 
                    
                    df_client["Base imposable au baréme"]  = ((((df_client["Salaire de base calcule"]+ df_client["Prime mensuelle (Barème) (DZD)"])-((df_client["Salaire de base calcule"]+ df_client["Prime mensuelle (Barème) (DZD)"])*0.09))+df_client["Indemnité 22jours"])/10)*10
                elif   df_client["Etablissement"].iloc[0] == "LG":
                        df_client["Base imposable au baréme"] = np.floor(((((df_client["Salaire de base calcule"] +df_client["Prime exeptionnelle (10%) (DZD)"] ) * 0.91)+ df_client["Indemnité 22jours"]))/ 10) * 10
                elif df_client["Etablissement"].iloc[0] == "G+D":
                    df_client["Base imposable au baréme"] = np.floor((((df_client["Salaire de base calcule"] + df_client["Prime exeptionnelle (10%) (DZD)"]) -df_client["Prime exeptionnelle (10%) (DZD)"]) * 0.91 + df_client["Indemnité 22jours"])/10)*10
                else:
                    df_client["Base imposable au baréme"] = np.floor((((df_client["Base cotisable"] - df_client["Prime exeptionnelle (10%) (DZD)"]- df_client["Indemnité non cotisable et imposable 10% (DZD)"]) * 0.91+ (df_client["Indemnité 22jours"])))/ 10) * 10
                    
                def irg_bareme(base):
                    b = np.ceil(base / 10) * 10  # PLAFOND(...;10) en Excel
                    
                    if b < 30000:
                        return 0
                    elif 30000 <= b <= 30860:
                        return (((( (b - 20000) * 0.23) - 1000) * 137/51) - (27925/8))
                    elif 30860 < b < 35000:
                        return (((( (b - 20000) * 0.23) * 0.60) * 137/51) - (27925/8))
                    elif 35000 <= b < 36310:
                        return (b - 20000) * 0.23 * 0.60
                    elif 36310 <= b < 40000:
                        return (b - 20000) * 0.23 - 1500
                    elif 40000 <= b < 80000:
                        return (b - 40000) * 0.27 + 3100
                    elif 80000 <= b < 160000:
                        return (b - 80000) * 0.30 + 13900
                    elif 160000 <= b < 320000:
                        return (b - 160000) * 0.33 + 37900
                    else:  # b >= 320000
                        return (b - 320000) * 0.35 + 90700

                df_client["IRG barème"] = df_client["Base imposable au baréme"].apply(irg_bareme)
                df_client["IRG 10%"] = df_client["Base imposable 10%"] * 0.10
                if df_client["Etablissement"].iloc[0] == "LG":
                    df_client["Salaire brut"] = (
                        (df_client["Salaire de base calcule"] +
                        df_client["Indemnité 22jours"]+
                        df_client["Indemnité non cotisable et imposable 10% (DZD)"]) 
                    )
                    # df_client["Salaire brut"] += np.where(
                    # df_client["Base de régul"] == "Salaire Brut", df_client["Régul"], 0)
                    df_client["Salaire net"] = ((df_client["Salaire de base calcule"]*0.91)+df_client["Indemnité 22jours"])+df_client["Base imposable 10%"]-df_client["IRG barème"]-df_client["IRG 10%"]
                    # df_client["Salaire net"] += np.where(
                    # df_client["Base de régul"] == "Salaire Net", df_client["Régul"], 0).round(0)
                elif df_client["Etablissement"].iloc[0] == "G+D":
                    df_client["Salaire brut"] = (
                        (df_client["Base cotisable"] +
                        (df_client["Indemnité 22jours"])+
                        df_client["Frais remboursement calcule"]) 
                    )
                    df_client["Salaire net"] = ((df_client["Salaire de base calcule"]+df_client["Indemnité 22jours"]+df_client["Prime exeptionnelle (10%) (DZD)"])-((df_client["Prime exeptionnelle (10%) (DZD)"]+df_client["Salaire de base calcule"])*0.09)-df_client["IRG barème"])
                else : 
                    df_client["Salaire brut"] = (
                        (df_client["Base cotisable"] +
                        (df_client["Indemnité 22jours"])+
                        df_client["Frais remboursement calcule"]) 
                    )
                    # df_client["Salaire brut"] += np.where(
                    # df_client["Base de régul"] == "Salaire Brut", df_client["Régul"], 0)
                    df_client["Salaire net"] = (
                        (df_client["Salaire brut"] -
                        df_client["Retenue CNAS employé"] -
                        df_client["IRG barème"] -
                        df_client["IRG 10%"]) 
                    )
                    # df_client["Salaire net"] += np.where(
                    # df_client["Base de régul"] == "Salaire Net", df_client["Régul"], 0).round(0)
                df_client["CNAS employeur"] = df_client["Base cotisable"] * 0.26
                df_client["Indemnités Non Cotisable - Mensuelle | Panier, Transport"] = df_client["Indemnité 22jours"]
                if  df_client["Etablissement"].iloc[0] == "Henkel":
                    df_client["Taxe formation et os"] = (df_client["Salaire de base calcule"] + df_client["Prime mensuelle (Barème) (DZD)"]+ df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] +df_client["Prime vestimentaire (DZD)"]) * 0.04
                    df_client["Cotisation œuvre sociale"] = 0
                    df_client["Taxe formation"] = 0
                elif df_client["Etablissement"].iloc[0] == "Maersk" :
                    df_client["Taxe formation et os"] = (df_client["Base cotisable"]) * 0.03
                    df_client["Cotisation œuvre sociale"] = 0
                    df_client["Taxe formation"] = 0
                else:
                    df_client["Cotisation œuvre sociale"] = df_client["Salaire brut"] * 0.02
                    df_client["Taxe formation"] = df_client["Salaire brut"] * 0.02
                    df_client["Taxe formation et os"] = 0
                    df_client["Masse salariale"] = (
                    df_client["Salaire brut"] +
                    df_client["CNAS employeur"] +
                    df_client["Cotisation œuvre sociale"] +
                    df_client["Taxe formation"]
                )
                df_client["Prime mensuelle"] = df_client["Prime mensuelle calcule"]
                if df_client["Etablissement"].iloc[0] == "Henkel":
                    df_client["Coût salaire"] = (
                        (df_client["Salaire net"]
                        + df_client["Taxe formation et os"]
                        + df_client["CNAS employeur"]
                        + df_client["IRG 10%"]
                        + df_client["IRG barème"]
                        + df_client["Retenue CNAS employé"]
                        )
                    )
                    # df_client["Coût salaire"] += np.where(
                    # df_client["Base de régul"] == "Cout salaire", df_client["Régul"], 0)
                    df_client["Coût congé payé"] = df_client.apply(calcul_cout_conge, axis=1)
                    fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                    
                    
                    if df_client["TAP"].iloc[0] == "Oui" :
                        df_client["TAP (DZD)"] = (df_client["Coût salaire"]+ df_client["Coût congé payé"] + df_client["Taux complément santé (DZD)"])*0.03
                        df_client["Facture HT"] = ((df_client["Coût salaire"] + df_client["Coût congé payé"]+ df_client["TAP (DZD)"] + df_client["Taux complément santé (DZD)"]) * fees_multiplicateur)
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                    else : 
                        df_client["TAP (DZD)"] = 0.0
                        df_client["Facture HT"] = ((df_client["Coût salaire"] + df_client["Coût congé payé"]+ df_client["TAP (DZD)"])* fees_multiplicateur)
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                      
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                elif df_client["Etablissement"].iloc[0] == "LG":
                    df_client["Coût salaire"] = (
                        (df_client["Salaire net"]
                        + df_client["Cotisation œuvre sociale"]
                        + df_client["Taxe formation"]
                        + df_client["CNAS employeur"]
                        + df_client["IRG 10%"]
                        + df_client["IRG barème"]
                        + df_client["Retenue CNAS employé"]
                       ) 
                    )
                    # df_client["Coût salaire"] += np.where(
                    # df_client["Base de régul"] == "Cout salaire", df_client["Régul"], 0)
                    df_client["Coût congé payé"] = df_client.apply(calcul_cout_conge, axis=1)
                    fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                    if df_client["TAP"].iloc[0] == "Oui" :
                        df_client["TAP (DZD)"] = (df_client["Coût salaire"] + ( df_client["Coût salaire"] * df_client["Fees etalent"])) * 0.02
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur) + df_client["TAP (DZD)"])+ df_client["Taux complément santé (DZD)"] 
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19
                    else : 
                        df_client["TAP (DZD)"] = 0.0
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur))+ df_client["Taux complément santé (DZD)"] 
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                elif df_client["Etablissement"].iloc[0] == "Maersk":
                    df_client["Coût salaire"] = (
                        (df_client["Salaire net"]
                        + df_client["Taxe formation et os"]
                        + df_client["CNAS employeur"]
                        + df_client["IRG 10%"]
                        + df_client["IRG barème"]
                        + df_client["Retenue CNAS employé"])
                    )
                    # df_client["Coût salaire"] += np.where(
                    # df_client["Base de régul"] == "Cout salaire", df_client["Régul"], 0)
                    df_client["Coût congé payé"] = df_client.apply(calcul_cout_conge, axis=1)
                    fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                    if df_client["TAP"].iloc[0] == "Oui" :
                        df_client["TAP (DZD)"] = (df_client["Coût salaire"] + ( df_client["Coût salaire"] * df_client["Fees etalent"])) * 0.02
                        df_client["Facture HT"] = ((df_client["Coût salaire"] + df_client["Coût congé payé"]+ df_client["TAP (DZD)"])* fees_multiplicateur)
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                    else : 
                        df_client["TAP (DZD)"] = 0.0
                        df_client["Facture HT"] = ((df_client["Coût salaire"] + df_client["Coût congé payé"]+ df_client["TAP (DZD)"])* fees_multiplicateur)
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19
                        
                elif df_client["Etablissement"].iloc[0] == "G+D":
                    df_client["Coût salaire"] = (df_client["Salaire de base calcule"] + df_client["Indemnité de panier calcule"] + df_client["Indemnité de transport calcule"] +df_client["Prime vestimentaire (DZD)"]+df_client["Frais remboursement calcule"]+df_client["Prime exeptionnelle (10%) (DZD)"]) 
                    # df_client["Coût salaire"] += np.where(
                    # df_client["Base de régul"] == "Cout salaire", df_client["Régul"], 0)
                    fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                    if df_client["TAP"].iloc[0] == "Oui" :
                        df_client["TAP (DZD)"] = (df_client["Coût salaire"] + ( df_client["Coût salaire"] * df_client["Fees etalent"])) * 0.02
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur) + df_client["TAP (DZD)"])
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                    else : 
                        df_client["TAP (DZD)"] = 0.0
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur))
                        df_client["Facture HT + NDF"] = df_client["Facture HT"]
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                else:
                    df_client["Coût congé payé"] = df_client.apply(calcul_cout_conge, axis=1)
                    fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                    df_client["Coût salaire"] = (
                        (df_client["Masse salariale"]
                        + df_client["Coût congé payé"]
                        + df_client["Taux complément santé (DZD)"]
                        ) 
                    )
                    # df_client["Coût salaire"] += np.where(
                    # df_client["Base de régul"] == "Cout salaire", df_client["Régul"], 0)
                    fees_multiplicateur = 1 + (df_client["Fees etalent"] / 100)
                    if df_client["TAP"].iloc[0] == "Oui" :
                        df_client["TAP (DZD)"] = (df_client["Coût salaire"] + ( df_client["Coût salaire"] * (df_client["Fees etalent"]/100))) * 0.02
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur) + df_client["TAP (DZD)"])
                        df_client["Facture HT + NDF"] = df_client["Facture HT"] 
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19

                    else : 
                        df_client["TAP (DZD)"] = 0.0
                        df_client["Facture HT"] = ((df_client["Coût salaire"] * fees_multiplicateur))
                        df_client["Facture HT + NDF"] = df_client["Facture HT"] 
                        # df_client["Facture HT + NDF"] = pd.to_numeric(df_client["Facture HT + NDF"], errors="coerce").fillna(0)
                        
                        if "df_rates" in st.session_state and not st.session_state.df_rates.empty:
                            df_rates = st.session_state.df_rates

                            # Récupérer les taux
                            euro_row = st.session_state.df_rates.loc[
                                st.session_state.df_rates["Devise"].str.contains("EUR"), "Achat"
                            ]

                            if not euro_row.empty:
                                euro_rate = euro_row.values[0]
                            else:
                                st.sidebar.warning("⚠️ Aucun taux EUR trouvé dans df_rates")
                                euro_rate = None

                            usd_rate  = df_rates.loc[df_rates["Devise"] == "USD", "Achat"].values[0]

                            # Choisir le bon taux en fonction de la sélection utilisateur
                            rate = euro_rate if st.session_state.devise_active == "EUR" else usd_rate
                            df_client["Facture HT en devise"] = df_client["Facture HT + NDF"] / rate
                            df_client["Facture TTC en devise"] = df_client["Facture HT en devis" ]*1.19
                # Construire un DataFrame avec toutes les nouvelles colonnes
                new_cols = pd.DataFrame({
                    "Frais remboursement": df_client["Frais remboursement calcule"],
                    "Salaire de base": df_client["Salaire de base calcule"],
                    "Indemnité de panier": df_client["Indemnité de panier calcule"],
                    "Indemnité de transport": df_client["Indemnité de transport calcule"],
                    "Facture TVA": df_client["Facture HT + NDF"] * (1.19 / 100),
                    "Facture TTC": df_client["Facture HT + NDF"] * (1 + 1.19 / 100),
                    "Observation": Observation
                })

                # Fusionner en une seule fois
                df_client = pd.concat([df_client, new_cols], axis=1).copy()

                print("👀 Employés bruts dans df_client :", df_client["Nom"].unique())

                st.write(df_client.head(50)) # On peut encapsuler ton code de calculs dans une fonction
                print(df_client[["Nom","N°"]])
                print(df_client.query("Nom == 'AMALOU'")[["Nom", "N°", "Mois"]])

                # 1. On définit les colonnes fixes (identité employé)
                # Colonnes fixes (identité employé)
                mois_ordre = [
                    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
                ]


                id_cols = ["Nom", "Prénom", "N°",  "Etablissement", "Année"]

                # Colonnes variables (toutes sauf identités + Mois)
                val_cols = [c for c in df_client.columns if c not in id_cols + ["Mois"]]

                # Pivot
                nb_candidats = df_client["N°"].nunique()

                if df_client["N°"].nunique() == 1:
                    # Un seul candidat → regrouper toutes les lignes par mois
                    employe_data = {}
                    for _, row in df_client.iterrows():
                        mois = row["Mois"]
                        for col in df_client.columns:
                            if col not in id_cols + ["Mois"]:
                                employe_data[f"{col}_{mois}"] = row[col]

                    # Réordonner les colonnes par mois_ordre
                    ordered_employe_data = {}
                    # Ajouter d'abord les colonnes statiques
                    for col in id_cols:
                        ordered_employe_data[col] = df_client.iloc[0][col]

                    # Ajouter les colonnes par mois
                    for mois in mois_ordre:
                        for col in val_cols:  # toutes les colonnes variables
                            key = f"{col}_{mois}"
                            if key in employe_data:
                                ordered_employe_data[key] = employe_data[key]

                    employe_data = ordered_employe_data
                else:
                   
                    df_pivot = (
                        df_client
                        .groupby(id_cols + ["Mois"], dropna=False)[val_cols]
                        .sum(min_count=1)
                        .unstack(fill_value=None)
                    )

                    # Aplatir MultiIndex colonnes
                    df_pivot.columns = [f"{val}_{mois}" for val, mois in df_pivot.columns]

                    # ✅ Forcer la présence de tous les employés
                    tous_employes = df_client[id_cols].drop_duplicates().set_index(id_cols)

                    df_pivot = (
                        df_pivot
                        .reindex(pd.MultiIndex.from_frame(tous_employes.reset_index()))
                        .reset_index()
                    )
                    




                    # Réordonner les colonnes par mois_ordre
                    colonnes_identite = id_cols
                    colonnes_mois = []
                    for mois in mois_ordre:
                        colonnes_mois.extend([c for c in df_pivot.columns if c.endswith(f"_{mois}")])

                    df_pivot = df_pivot[colonnes_identite + colonnes_mois]

                    print(df_pivot[["Nom","N°"]])

                # 📥 Génération et téléchargement Excel
                # --------------------------------------
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_client.to_excel(writer, index=False, sheet_name='Calculs')
                    workbook = writer.book
                    worksheet = writer.sheets['Calculs']

                    # Style entête
                    header_format = workbook.add_format({
                        'bold': True, 'text_wrap': True, 'valign': 'middle',
                        'align': 'center', 'fg_color': '#0a5275',
                        'font_color': 'white', 'border': 1
                    })
                    for col_num, value in enumerate(df_client.columns.values):
                        worksheet.write(0, col_num, value, header_format)

                    # Ajuster largeur colonnes
                    for i, col in enumerate(df_client.columns):
                        col_width = max(df_client[col].astype(str).map(len).max(), len(col)) + 2
                        worksheet.set_column(i, i, col_width)

                st.download_button(
                    label="📊 Télécharger les résultats en Excel",
                    data=output.getvalue(),
                    file_name=f"{st.session_state.selected_client}_calculs.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

               
                # 📥 Génération et téléchargement PDF par employé
                # ------------------------------------------------
                st.markdown("### 📥 Télécharger la facture PDF par employé")

                # 📌 Nouveau fichier par client
                
                # Nom du client
                client_name = str(df_client["Etablissement"].iloc[0]).strip()
                fichier_client = f"{client_name}_factures.xlsx"

                wb = Workbook()
                wb.remove(wb.active)  # supprimer la feuille vide

                if nb_candidats == 1:
                    # Un seul employé → créer UNE seule feuille consolidée
                    nom = df_client.iloc[0]["Nom"].replace(" ", "_")
                    matricule = str(df_client.iloc[0]["N°"])
                    
                    ws = wb.create_sheet(title=f"{matricule}_{nom}")
                    generer_facture_excel_sheet(employe_data, ws, wb)
                else:
                    # Plusieurs employés → plusieurs feuilles
                    for idx, row in df_pivot.iterrows():
                        nom = str(row.get("Nom", f"employe_{idx}")).strip().replace(" ", "_")
                        matricule = str(row.get("N°", f"id_{idx}")).strip()

                        ws = wb.create_sheet(title=f"{matricule}_{nom}")
                        generer_facture_excel_sheet(row.to_dict(), ws, wb)

                # Sauvegarde
                wb.save(fichier_client)

                # Télécharger
                with open(fichier_client, "rb") as f:
                    st.download_button(
                        label=f"📘 Télécharger factures client {client_name}",
                        data=f.read(),
                        file_name=fichier_client,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                drive_file_id = upload_to_drive(
                    fichier_client,  # <-- utiliser fichier_client ici
                    client_name=row.get("Etablissement", "Inconnu") if nb_candidats > 1 else client_name,
                    root_folder_id="0AM1AktJToIM1Uk9PVA"
                )

                # Supprimer le fichier local
                os.remove(fichier_client)

            else:
                st.warning("⚠️ Aucun employé trouvé pour ce client ")
        else:
            st.info("Veuillez d'abord téléverser le fichier récapitulatif global dans la barre latérale.")
            


